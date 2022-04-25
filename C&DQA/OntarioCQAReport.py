import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side, Font, Alignment
import gettingFeeCodes
import CQAUtilities
import Utilities
from Utilities import *

def OntarioQuebecCQA(workbook, CQAREF):
    print  'ONTARIO /QEUEBEC REPORT', workbook, CQAREF
    grey_highlight = PatternFill(start_color='E6E6E3', end_color='E6E6E3', fill_type='solid')
    sheet = workbook.get_sheet_by_name("CCME (Ontario)")
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")
    # sprint('CULMN-NUM, NAME, VALUE, ROW-INDX ')

    array_values, _, _ = CQAUtilities.OntarioResults(CQAREF)
    Utilities.round_all_array_values(array_values)
    item_dict = Utilities.getValuesForAGIndex(CQAREF)
    pe_m3_dict = CQAUtilities.getOtherResults(CQAREF)
    k_m3_value, mg_m3_value, ca_m3_value, na = andres_percent_calc(andres_cec_calc(get_cec_values(CQAREF)), get_cec_values(CQAREF))
    for i in array_values:
        print(i)

    # A.
    sheet.cell(row=9, column=4).value = array_values[0][2]  # Arsenic
    sheet.cell(row=10, column=4).value = array_values[1][2]  # Cadmium
    sheet.cell(row=11, column=4).value = array_values[2][2]  # chromium
    sheet.cell(row=12, column=4).value = array_values[3][2]  # cobalt
    sheet.cell(row=13, column=4).value = array_values[4][2]
    sheet.cell(row=14, column=4).value = array_values[5][2]
    sheet.cell(row=15, column=4).value = array_values[6][2]
    sheet.cell(row=16, column=4).value = array_values[7][2]
    sheet.cell(row=17, column=4).value = array_values[8][2]
    sheet.cell(row=18, column=4).value = array_values[9][2]
    sheet.cell(row=19, column=4).value = array_values[10][2]

    # B.
    sheet.cell(row=24, column=4).value = Utilities.BDL_PERCENT_check(array_values[29][2])
    sheet.cell(row=25, column=4).value = Utilities.BDL_PERCENT_check(array_values[30][2])
    sheet.cell(row=26, column=4).value = Utilities.BDL_PERCENT_check(array_values[11][2])
    sheet.cell(row=28, column=4).value = Utilities.BDL_PERCENT_check(array_values[12][2])
    sheet.cell(row=29, column=4).value = Utilities.BDL_PERCENT_check(array_values[13][2])

    # C.
    sheet.cell(row=33, column=4).value = Utilities.BDL_PERCENT_check(array_values[14][2])
    # sheet.cell(row=34, column=4).value = array_values[][]

    # D.
    sheet.cell(row=40, column=4).value = array_values[15][2]
    sheet.cell(row=41, column=4).value = array_values[16][2]

    # E.
    sheet.cell(row=46, column=6).value = str(array_values[17][2]) + '%'
    sheet.cell(row=47, column=6).value = str(array_values[18][2]) + '%'

    sheet.cell(row=53, column=6).value = array_values[19][2]
    sheet.cell(row=54, column=6).value = array_values[20][2]
    sheet.cell(row=55, column=6).value = CQAUtilities.get_partcile(CQAREF)
    sheet.cell(row=56, column=6).value = "{:.1f}".format(round(float(pe_m3_dict['salt']), 1))  # salt
    sheet.cell(row=57, column=6).value = "{:.2f}%".format(na)
    sheet.cell(row=59, column=6).value = str(k_m3_value) + '%'  # perk
    sheet.cell(row=60, column=6).value = str(mg_m3_value) + '%'  # perma
    sheet.cell(row=61, column=6).value = str(ca_m3_value) + '%'  # perca

    # APENDIX 3
    sheet.cell(row=100, column=4).value = str(CQAUtilities.get_dry_matter(CQAREF)) + '%'
    sheet.cell(row=101, column=4).value = item_dict['PH']
    sheet.cell(row=102, column=4).value = "{:.0f}".format(float(array_values[22][2])) # remove decimal
    sheet.cell(row=103, column=4).value = array_values[20][2]

    # FERTILIZER

    sheet.cell(row=105, column=4).value = str(round(float(Utilities.getNitrogen(CQAREF)), 2)) + '%'
    sheet.cell(row=106, column=4).value = array_values[23][2]
    sheet.cell(row=107, column=4).value = str(array_values[24][2]) + '%'
    sheet.cell(row=108, column=4).value = str(array_values[25][2]) + '%'
    sheet.cell(row=109, column=4).value = str(array_values[26][2]) + '%'
    sheet.cell(row=110, column=4).value = "{:.2f}%".format(array_values[27][2])
    sheet.cell(row=111, column=4).value = array_values[28][2]

    # AGINDEX----------------------------------------------------------
    item_dict = Utilities.getValuesForAGIndex(CQAREF)

    print('getting agindex values')
    DryMatter = float(CQAUtilities.get_dry_matter(CQAREF))
    Nitrogen = float(Utilities.getNitrogen(CQAREF))  # stand in for real value
    Phosphorus = (float(CQAUtilities.get_Agindex_Phosphorus(CQAREF) * (DryMatter / 100) / 10000)) * 2.2914
    Potassium = (float(CQAUtilities.get_Agindex_Potassium(CQAREF) * (DryMatter / 100) / 10000)) * 1.2046
    Sodium = float(CQAUtilities.get_Agindex_Sodium(CQAREF) * (DryMatter / 100))
    Chloride = float(item_dict['CL']) / 10000
    # print('dude wtf', Chloride, type(Chloride), type(float(Chloride)))

    print('PHOSPHORUS:', Phosphorus)  # *
    print('Potassium:', Potassium)  # *
    print('Nitrogen:', Nitrogen)
    print('sodium:', Sodium)  # *
    print('DryMatter:', DryMatter)
    print('Chloride:', Chloride)
    value1 = (Nitrogen + Phosphorus + Potassium)
    value2 = Sodium + (Chloride)
    print 'LEFTSIDE:', value1
    print 'RIGHTSIDE:', value2
    ag_index = value1 / value2
    print('AGINDEX = ', ag_index)
    sheet.cell(row=113, column=4).value = round(ag_index, 2)
    sheet.cell(row=113, column=6).value = CQAUtilities.agindex_text(ag_index)

    # -------- FORMATTING
    CQAUtilities.CQA_ONT_FORMATTING(sheet)

    # --- Removing or Adding Percent Signs
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')
    # ag_index_jpg = Image('C:\CQA\FULL CQA - DQA\C&DQA\Photos\Agindex.jpg')
    ag_index_png = Image('C:/CQA\FULL CQA - DQA/C&DQA/Photos/agindex.png')
    sheet.add_image(ag_index_png, 'A115')

    # HIGHLIGHTING
    import HighlighterChecker
    HighlighterChecker.get_ontario_cqa_constraints_A(sheet)

    # REMOVING THE PLACES WHERE BDL HAS PERCENT OR NA
    CQAUtilities.remove_BDL_percent(sheet)

    #AGINDEX FONT
    font_black = Font(color='000000', size=10)
    sheet.cell(row=113, column=6).font = font_black
    current_cell = sheet['F113']
    current_cell.alignment = current_cell.alignment.copy(wrapText=True)

    # ------
    # putting in the images------------------------------------
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')
    img = Image('al.jpg')
    sheet.add_image(img, 'A1')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'H1')

    img = Image('al.jpg')
    sheet.add_image(img, 'A94')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'H94')

    # ----------------------------------------------------------------
    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport",
                                CQAREF)

    filename = saveLocation + "\%sReport.xlsx" % CQAREF
    Workbook.save(workbook, filename)

    # template path | C:\CQA\FULL CQA - DQA\C&DQA\Templates\TEMPLATE ON WRITTEN
