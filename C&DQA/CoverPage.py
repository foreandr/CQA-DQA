import time, os
import mysql.connector
from mysql.connector import errorcode
from openpyxl.styles import Border, Side
from Utilities import FixFormatting
import shutil, os, sys
import Colors
import SQL_CONNECTOR

def coverPageWrite(CQARef, workingFolder):
    print Colors.bcolors.OKCYAN + "\nExecuting Cover Page Write" + Colors.bcolors.ENDC

    # Its the template file and where the file will be saved (The working folder + the name of the file it will spit out
    templateFile = r'C:\CQA\FULL CQA - DQA\C&DQA\Templates\cover.xlsx'
    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport", CQARef)
    cqaFile = saveLocation + r'\%sCover.xlsx' % (CQARef)

    # The dictonary for all the cover items
    coverDict = {
        'A1': None,
        'A2': None,
        'A3': None,
        'A4': None,
        'A5': None,
        'A6': None,
        'A7': None,
        'A8': None,
        'A9': None,
        'A10': None,
        'A111': None
    }

    # print 'GETTING SQL CONNECTION'
    cnx = SQL_CONNECTOR.test_connection()

    # --------A1 Querry--------#
    # Finds the company name
    cursor = cnx.cursor()
    query = """SELECT company FROM alms.report WHERE refno='%s'and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        company = str(item[0])
    coverDict['A1'] = company

    # --------A2 Querry--------#
    # Finds the company address
    cursor = cnx.cursor()
    query = """SELECT address1 FROM alms.report WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        address = str(item[0])
    coverDict['A2'] = address

    # ------city Querry--------#
    # finds the city
    cursor = cnx.cursor()
    query = """SELECT city FROM alms.report WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        city = str(item[0])

    # ------State Querry-------#
    # finds the state for sale purposes
    cursor = cnx.cursor()
    query = """SELECT state FROM alms.report WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        state = str(item[0])
    coverDict['A111'] = state

    # Convert State name ------ Want to change the abbreviated province name to full name
    print Colors.bcolors.HEADER + Colors.bcolors.UNDERLINE + '-----------------------------PROVINCES------------------------' + Colors.bcolors.ENDC
    query = """SELECT name FROM (alms.report INNER JOIN alms.provinces ON report.state = provinces.abbreviation) WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        print Colors.bcolors.OKBLUE + "State: " + str(item)
        state = item[0].encode('utf-8').strip() # VERY IMPORTANT
        coverDict['A111'] = state

    # finds the zip code
    query = """SELECT zip FROM alms.report WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        zipC = str(item[0])

    # creates the whole address using the before created variables
    address = city + ', ' + state + ' ' + zipC
    coverDict['A3'] = address

    # --------A4 Querry---------#
    # Creates the attention
    cursor = cnx.cursor()
    query = """SELECT attn FROM alms.report WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        attention = str(item[0])
    coverDict['A4'] = attention

    # --------A5 and A6 Querry------#
    # Gets the report numbers
    cursor = cnx.cursor()
    query = "SELECT * FROM alms.report WHERE refno='%s'" % (CQARef)
    cursor.execute(query)
    reportNumbers = []
    # Stores repost numbers in report list
    for item in cursor:
        tempList = [item[0], item[1]]
        reportNumbers.append(tempList)

    # Sortes out the report numbers into soil report number or enviromental
    soilReport = None
    envReport = None
    for item in reportNumbers:
        if str(item[1]) == "SOIL":
            soilReport = str(item[0])
            coverDict['A5'] = soilReport
        elif str(item[1]) == "ENVI":
            envReport = str(item[0])
            coverDict['A6'] = envReport  # this is where we add A7 for the other envrio report
        else:
            print "error"

    # -----A8 Querry-----#
    # Gets the sample ID
    cursor = cnx.cursor()
    query = """SELECT grow_1 FROM alms.report WHERE refno='%s' and module='SOIL' """ % CQARef
    cursor.execute(query)
    for item in cursor:
        sampleID = str(item[0])
    print Colors.bcolors.OKBLUE + "Sample ID:" + sampleID + Colors.bcolors.ENDC
    coverDict['A8'] = sampleID

    # ---A9 Querry---#
    # Gets the Current date
    currentDate = time.localtime()[0:3]
    coverDict['A9'] = r'%d-%d-%d' % (currentDate[0], currentDate[1], currentDate[2])

    '''
    # ---A10 Querry---#
    catagory = findCoverCatagory(CQARef)
    if catagory == 'A' or catagory == 'B' or catagory == 'AA':
        print 'sweet stuff'
        coverDict['A10'] = 'Category %s' % (catagory)
    else:
        print 'the right one'
        coverDict['A10'] = 'Exceeds Guidelines'
    '''

    # ---B18 Query FEEDSTOCK---#
    print 'executing query 18 for feedstock'
    cursor = cnx.cursor()
    query = """
        SELECT report.refno, feedstock.description
        FROM (alms.env_samp env_samp
        INNER JOIN alms.feedstock feedstock
        ON (env_samp.feedstock_code = feedstock.code))
        INNER JOIN alms.report report ON (env_samp.rptno = report.rptno)
        WHERE (report.refno = '%s')"""% CQARef
    wanted_description = "Not Specified"
    cursor.execute(query)
    for item in cursor:
        wanted_description = item[1] #  second item in tuple is description, first is iD
        print wanted_description
        print Colors.bcolors.UNDERLINE + Colors.bcolors.OKGREEN + "[REFNO: " + str(CQARef) + "] CURRENT FEEDSTOCK DESC: " + str(item) + Colors.bcolors.ENDC


    # fill in the excel sheet by checking each box that should have a number in it
    import openpyxl
    wb = openpyxl.load_workbook(templateFile)
    if wb is None:
        print "Invalid Workbook"
    sheet = wb.get_sheet_by_name('CoverPage')

    # FEEDSTOCK PLACEMENT # Works fine
    sheet.cell(row=18, column=2).value = wanted_description

    # A1-A3 Address
    for i in range(7, 10):
        value = sheet.cell(row=i, column=2).value
        cellName = str(value)
        valueToFill = coverDict[cellName]
        sheet.cell(row=i, column=2).value = valueToFill

    # A111 Province name
    value = sheet.cell(row=17, column=5).value
    cellName = str(value)
    valueToFill = coverDict[cellName]
    sheet.cell(row=17, column=5).value = valueToFill

    # A4 Attention
    value = sheet.cell(row=11, column=2).value
    cellName = str(value)
    valueToFill = coverDict[cellName]
    sheet.cell(row=11, column=2).value = valueToFill

    # A5-A7 report numbers
    for i in range(13, 16):
        value = sheet.cell(row=i, column=2).value
        cellName = str(value)
        valueToFill = coverDict[cellName]
        sheet.cell(row=i, column=2).value = valueToFill

    # A8  Sample ID in 2 cells
    value = sheet.cell(row=10, column=8).value
    cellName = str(value)
    try:
        valueToFill = coverDict[cellName]
    except KeyError:
        valueToFill = 'Error'
    sheet.cell(row=10, column=8).value = valueToFill
    sheet.cell(row=22, column=1).value = valueToFill

    # A9 Reported Date
    value = sheet.cell(row=14, column=8).value
    cellName = str(value)
    valueToFill = coverDict[cellName]
    sheet.cell(row=14, column=8).value = valueToFill

    # A10 category and pass fail comment
    value = sheet.cell(row=22, column=3).value
    cellName = str(value)
    valueToFill = coverDict[cellName]
    sheet.cell(row=22, column=3).value = valueToFill

    # Creates the border styles for the different side types
    thick = Side(border_style="thick", color='00B050')
    thin = Side(border_style="medium", color='00B050')

    # Sets the border
    border = Border(right=thick, bottom=thick)
    # Calls the FixFormatting function
    FixFormatting(sheet, 'A21:B23', border)

    border = Border(bottom=thick)
    FixFormatting(sheet, 'C21:I23', border)

    border = Border(right=thick)
    FixFormatting(sheet, 'A24:B27', border)

    border = Border(top=thin, left=thin, right=thin, bottom=thick)
    FixFormatting(sheet, 'A21:I27', border)

    # lines under the signatures
    thinnish = Side(border_style="thin", color='000000')
    border = Border(bottom=thinnish)
    FixFormatting(sheet, 'A36:D36', border)
    border = Border(bottom=thinnish)
    FixFormatting(sheet, 'F36:I36', border)

    # imports all the photos to the sheet
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\NewFormatCQA\Photos')
    img = Image('coverCp.png')
    sheet.add_image(img, 'D1')
    img = Image('hs.jpg')
    sheet.add_image(img, 'B35')
    img = Image('ian.bmp')
    sheet.add_image(img, 'H35')
    img = Image('alcover.png')
    sheet.add_image(img, 'A40')
    img = Image('cpLogCover.png')
    sheet.add_image(img, 'G40')

    # save file
    wb.save(cqaFile)
    wb = None
    # close the cursor
    cursor.close()
    cnx.close()
