import os, sys, shutil
import mysql.connector
from mysql.connector import errorcode
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import Colors
import SQL_CONNECTOR
import Utilities
from OntarioPart import slicePercentOffUnicode, decisionAgIndex
from Utilities import FixFormatting, config, formatDict, organicCarbon, calculateCEC, percentCalc, merge_two_dicts


def CCMEPrint(Workbook, finalResult, CQARef):
    highlightList = FailStandard(CQARef)
    # This is the sheet name inside of the template must be exact
    sheet = Workbook.get_sheet_by_name("CCME (Provinces&Territories)")
    # Sets the colour of the highlight/fill to highlight the failed values
    highlight = PatternFill(start_color='F3F315', end_color='F3F315', fill_type='solid')

    # Unicode to add subscripts
    sheet['A34'] = u'CO\u2082 Respiration Rate'
    sheet['A35'] = u'CO\u2082 Respiration Rate'
    sheet['A36'] = u'O\u2082 Uptake Respiration Rate'
    sheet['A37'] = u'O\u2082 Uptake Respiration Rate'

    # Prints the value to the excel sheet for each one
    print "----Trace Elements----"
    # 8-18: 1-11
    for i in range(10, 21, 1):
        # finds what value is currently in the cell(A1 for example)
        value = sheet.cell(row=i, column=4).value
        # sheet.cell(row=i, column=4).fill = highlight # JUST TESTING
        try:
            # puts the value of the cell into string format
            cellName = str(value)
            # uses the value to search for the result
            valueToFill = finalResult[cellName]
        except KeyError:
            # If there isnt a result fill in a error
            valueToFill = "Error"

        sheet.cell(row=i, column=4).value = valueToFill
        if highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight

    print "----foreign matter----"
    # D28: 12
    value = sheet.cell(row=26, column=4).value
    try:
        cellName = str(value)
        valueToFill = finalResult[cellName]
    except KeyError:
        valueToFill = "Error"

    sheet.cell(row=26, column=4).value = valueToFill
    if highlightList[str(value)] == '1':
        sheet.cell(row=26, column=4).fill = highlight

    # D30-31: 13-14
    for i in range(29, 31, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=4).value = valueToFill
        if highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight

    print "----Maturity/Stability----"
    # D35: 15
    value = sheet.cell(row=34, column=4).value
    try:
        cellName = str(value)
        valueToFill = finalResult[cellName]
    except KeyError:
        valueToFill = "Error"

    sheet.cell(row=34, column=4).value = valueToFill
    if highlightList[str(value)] == '1':
        sheet.cell(row=34, column=4).fill = highlight

    print"----Pathogens----"
    # D44-45: 16-17
    for i in range(41, 43, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=4).value = valueToFill
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight

    print"----CFIA----"
    # F50-51: 18-19
    for i in range(47, 49, 1):
        value = sheet.cell(row=i, column=6).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=6).value = valueToFill

    print"----Finished Compost Quality----"
    # f62-66: 20-24
    for i in range(55, 60, 1):
        value = sheet.cell(row=i, column=6).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=6).value = valueToFill

    # f68-70: 25-27
    for i in range(61, 64, 1):
        value = sheet.cell(row=i, column=6).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=6).value = valueToFill

    print"----Compost Agricultural Product Value----"
    # D98-101: 28-31
    for i in range(98, 101, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=4).value = valueToFill

    # D103-109: 32-38
    for i in range(103, 109, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=4).value = valueToFill

    print finalResult

    # manual additions
    sheet.cell(row=101, column=4).value = finalResult['31']
    sheet.cell(row=109, column=4).value = finalResult['38']
    # print ' TESTING FINAL RESULT ' + str(finalResult['31'])
    # print ' TESTING FINAL RESULT ' + str(finalResult['38'])

    # Below is all the code used to fix the borders by giving it a range of cell values(eg. A10:I10) and using Border to change the thickness size
    # Creates the border styles for the different side types
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    # Trace metals
    # This line sets the border values
    border = Border(top=thick, left=thick, right=thick, bottom=thick)
    # This line calls the FixFormatting function
    FixFormatting(sheet, 'B7:D7', border)
    # This changes the border for the one needed and it just kinda keeps rotating
    border = Border(bottom=thick)
    FixFormatting(sheet, 'B20:D20', border)
    border = Border(right=thick)
    FixFormatting(sheet, 'E8:H8', border)
    FixFormatting(sheet, 'E9:H9', border)
    border = Border(bottom=thick, top=thick, right=thick)
    FixFormatting(sheet, 'F7:H7', border)
    border = Border(bottom=thick)
    FixFormatting(sheet, 'E9:H9', border)
    border = Border(bottom=thick, right=thick)
    FixFormatting(sheet, 'F20:H20', border)
    border = Border(top=thick, left=thick, right=thick, bottom=thick)
    FixFormatting(sheet, 'B7:C7', border)

    # This is for the walls of the borders where i can just run a for loop instead of doing it manually
    border = Border(right=thin, bottom=thin)
    for i in range(10, 20):
        cellNumber = 'B%s:C%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    border = Border(bottom=thin, right=thick)
    for i in range(10, 20):
        cellNumber = 'F%s:H%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # Foreign Matter
    # the boder isn't working on bottom of D26 fixed first line
    border = Border(bottom=thin)
    FixFormatting(sheet, 'D26:D27', border)
    border = Border(bottom=thick, right=thick)
    FixFormatting(sheet, 'H30:I30', border)
    border = Border(top=thick, left=thin, right=thick, bottom=thick)
    FixFormatting(sheet, 'A24:I24', border)
    border = Border(left=thick)
    FixFormatting(sheet, 'A24:A30', border)
    border = Border(top=thick, left=thin, right=thin, bottom=thin)
    FixFormatting(sheet, 'E25:F27', border)
    border = Border(top=thick, left=thin, right=thick, bottom=thin)
    FixFormatting(sheet, 'G25:I27', border)
    border = Border(top=thin, left=thin, right=thin, bottom=thick)
    FixFormatting(sheet, 'E28:F30', border)
    border = Border(bottom=thick)
    FixFormatting(sheet, 'A30:C30', border)

    border = Border(right=thin, bottom=thin, left=thin)
    for i in range(24, 30):
        cellNumber = 'B%s:C%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    border = Border(right=thick)
    for i in range(24, 30):
        cellNumber = 'H%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # Maturity and stability
    border = Border(top=thick, bottom=thick)
    FixFormatting(sheet, 'A33:I33', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A33:I37', border)
    border = Border(bottom=thick)
    FixFormatting(sheet, 'A37:C37', border)
    # this bottom border should be thick but isn't

    border = Border(right=thin, bottom=thin)
    for i in range(34, 37, 2):
        cellNumber = 'A%s:C%s' % (i, i + 1)
        FixFormatting(sheet, cellNumber, border)

    border = Border(left=thin, bottom=thin, right=thin)
    for i in range(34, 37, 2):
        cellNumber = 'E%s:I%s' % (i, i + 1)
        FixFormatting(sheet, cellNumber, border)

    # Pathogens
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'A40:I40', border)
    border = Border(right=thick, bottom=thin)
    FixFormatting(sheet, 'A41:I41', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A42:I42', border)

    # CFIA
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'C46:G46', border)
    border = Border(right=thick, bottom=thin)
    FixFormatting(sheet, 'C47:G47', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'C48:G48', border)

    # Compost Quality
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'C54:G54', border)

    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'C63:G63', border)

    border = Border(right=thick, bottom=thin)
    for i in range(54, 63):
        cellNumber = 'C%s:G%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # reference Compost Quality
    border = Border(bottom=thick, top=thick)
    FixFormatting(sheet, 'A67:I67', border)

    border = Border(bottom=thick, top=thick)
    FixFormatting(sheet, 'A72:I72', border)

    border = Border(bottom=thin, top=thin)
    for i in range(73, 82):
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    border = Border(bottom=thick)
    FixFormatting(sheet, 'A82:I82', border)

    border = Border(top=thin)
    FixFormatting(sheet, 'A83:I83', border)

    border = Border(top=thick, right=thick, bottom=thick)
    FixFormatting(sheet, 'A96:I96', border)

    border = Border(top=thin, right=thick, bottom=thin)
    for i in range(98, 110):
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    border = Border(bottom=thick, top=thick)
    FixFormatting(sheet, 'A111:I111', border)

    '''DOING CALCULATION FOR AGINDEX-------------------------------------------------------------------------------------------'''
    value = 'TEMP'
    N = sheet.cell(row=103, column=4).value  # TOTAL NITROGEN
    P2O5 = sheet.cell(row=105, column=4).value  # PHOSPHATE
    K20 = sheet.cell(row=106, column=4).value  # POTASSIUM

    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = "select rep.rptno, rep.refno, env.feecode, env.result from alms.report rep inner join alms.env_data env on rep.rptno = env.rptno where env.feecode = 'mnac380' and rep.refno = '%s' " % CQARef
    cursor.execute(query)
    NA = ''
    for i in cursor:
        NA = i[3]

    cursor = cnx.cursor()
    CL = ''
    query = "select soil.cl from alms.report rep inner join alms.soil soil on rep.rptno = soil.rptno where rep.refno = '%s'" % CQARef
    cursor.execute(query)
    for i in cursor:
        # print i
        CL = i[0]

    initialUnicodeList = [N, P2O5, K20, NA, CL]
    for i in initialUnicodeList:
        print i
    updatedFloatList = []
    for i in initialUnicodeList:
        # i = str(i)
        print 'printing initial float list'
        print i
        j = slicePercentOffUnicode(str(i))
        updatedFloatList.append(j)

    for i in updatedFloatList:
        # print i
        # print type(i)
        pass

    drymatter = sheet.cell(row=98, column=4).value  # TOTAL NITROGEN

    # (dry matter /100 )* sodium
    Nitrogen = updatedFloatList[0]
    Phosphorus = updatedFloatList[1]
    Potassium = updatedFloatList[2]
    Sodium = updatedFloatList[3]
    Chloride = updatedFloatList[4]
    DryMatter = slicePercentOffUnicode(str(drymatter))

    a_index = (Nitrogen + Phosphorus + Potassium) / ((Sodium * (DryMatter / 100)) + (Chloride / 10000))
    sheet.cell(row=111, column=4).value = round(a_index)  # cast to integer

    # Putting in the comment
    print decisionAgIndex(a_index)
    font_c = Font(color='000000', size=10)
    sheet.cell(row=111, column=6).value = decisionAgIndex(a_index)
    sheet.cell(row=111, column=6).font = font_c

    # AGINDEX ADDITION
    from openpyxl.drawing.image import Image
    img = Image("K:/2022Ontario/Student2022/ANDRE-CQA 2 report system/Photos/agindex.png")
    sheet.add_image(img, 'b113')

    #font_c = Font(color='000000', size=10)
    # sheet.cell(row=115, column=1).value = "salt injury probably"
    # sheet.cell(row=115, column=3).value = "apply on salts with excellent drainage chracteristics good water quality and low salts"
    # sheet.cell(row=115, column=6).value = "apply on salts with poor drainage poor water quality, or high salts"
    # sheet.cell(row=115, column=9).value = "for all salts"
    for i in range(1, 10):  # Changing color from red to black
        sheet.cell(row=115, column=i).font = font_c

    #sheet.cell(row=116, column=5).value = "fig 05.02-F1 AgIndex interpretation and use for commmon edaphic conditions"
    #sheet.cell(row=116, column=5).font = font_c

    # imports all the photos to the sheet
    from openpyxl.drawing.image import Image
    # The directory of where the photos are kept
    os.chdir(r'C:\CQA\NewFormatCQA\Photos')
    # AL
    # Creating an image variable
    img = Image('al.jpg')
    # Adding it to the sheet at A1
    sheet.add_image(img, 'A1')
    img2 = Image('al.jpg')
    sheet.add_image(img2, 'A50')
    img3 = Image('al.jpg')
    sheet.add_image(img3, 'A93')

    # Cp
    img = Image('cp.png')
    sheet.add_image(img, 'H1')
    img = Image('cp.png')
    sheet.add_image(img, 'H50')
    img = Image('cp.png')
    sheet.add_image(img, 'H93')

    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport", CQARef)
    Workbook.save(saveLocation + "\%sReport.xlsx" % (CQARef))


def findCCMECatagory(CQARef):
    """makes a list and adds the values to the list and checks what category they're in, returns the category letter"""
    # gets the results from the CCME Results Function
    ENVResult = CCMEResults(CQARef)
    # Makes a dictionary for the results needed(1-17 are needed)
    results = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # Puts the 17 results needed into a separate list called valueCatList
    valueCatList = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # The max values for each category(Arsenic catA=13 catB=75) -1 means there is no value in that category and will be skipped
    catA = {
        '1': 13,
        '2': 3,
        '3': 210,
        '4': 34,
        '5': 400,
        '6': 150,
        '7': 0.8,
        '8': 5,
        '9': 62,
        '10': 2,
        '11': 700,
        '12': 1,
        '13': 0,
        '14': 0,
        '15': 4,
        '16': 1000,
        '17': 3
    }

    catB = {
        '1': 75,
        '2': 20,
        '3': -1,
        '4': 150,
        '5': -1,
        '6': 500,
        '7': 5,
        '8': 20,
        '9': 180,
        '10': 14,
        '11': 1850,
        '12': 2,
        '13': 3,
        '14': 0,
        '15': -1,
        '16': -1,
        '17': -1
    }

    # The category defaults to A
    finalResult = 'A'

    # Stores the values into valueCatList
    for i in results.keys():
        valueCatList[i] = ENVResult[i]

    for key in results.keys():
        # Temp is used to store and modify the current value its on to remove percent signs
        temp = None
        try:
            temp = valueCatList[key].replace('%', '')

        # If the value is an integer then dont bother doing it
        except AttributeError:
            temp = valueCatList[key]

        # Make sure that all the values that can be numbers are numbers
        try:
            value = float(temp)

        # If it can't be a number then don't change anything
        except ValueError:
            value = valueCatList[key]

        # If the value is BDL then it has to be A
        if value == 'BDL*':
            results[key] = 'A'

        # If the value is a string then find out what one it is and find the category by its type
        elif type(value) == str:
            # If the value of salmonella is negative then its A
            if value == 'Negative' and key == '17' or value == 'NEGATIVE' and key == '17':
                results[key] = 'A'
            # If the value of salmonella is positive than it exceeds
            elif value == 'Positive' and key == '17' or value == 'POSITIVE' and key == '17':
                results[key] = 'Exceeds'
            # If its Fecal Coliform and its <3 then its A
            elif value == '<3' and key == '16':
                results[key] = 'A'
            # if its Fecal Coliform and its >1000 then it exceeds
            elif value == '>1000' and key == '16':
                results[key] = 'Exceeds'
            # if its fecal and its <3 BUT also salmonella is positive than it exceeds
            elif value == '<3' and key == '16' and value == 'Positive' and key == '17':
                results[key] = 'Exceeds'

        # If the value is less then category A then its A
        elif value <= catA[key]:
            results[key] = 'A'

        elif value > catA[key]:
            # If its above category A and Category B doesn't exist then it Exceeds
            if catB[key] == -1:
                results[key] = 'Exceeds'

            else:
                # If the value is below category B then its B
                if value <= catB[key]:
                    results[key] = 'B'
                # If it doesn't fit in any of the above things then it Exceeds
                else:
                    results[key] = 'Exceeds'

    # goes through all the dictionary results and finds the highest category letter and returns that one
    for key in results.keys():
        parameter = results[key]

        if finalResult == parameter:
            finalResult = parameter
        elif finalResult == 'A' and parameter == 'B':
            finalResult = parameter
        elif (finalResult == 'B' or finalResult == 'A') and parameter == 'Exceeds':
            finalResult = parameter

    return finalResult


def FailStandard(CQARef):
    # gets the results from the CCME Results Function
    ENVResult = CCMEResults(CQARef)
    # Makes a dictionary for the results needed(1-17 are needed)
    results = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # Puts the 17 results needed into a separate list called valueCatList
    valueCatList = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    catMax = {
        '1': 75,
        '2': 20,
        '3': 210,
        '4': 150,
        '5': 400,
        '6': 500,
        '7': 5,
        '8': 20,
        '9': 180,
        '10': 14,
        '11': 1850,
        '12': 2,
        '13': 3,
        '14': 1,
        '15': 4,
        '16': 1000,
        '17': 3
    }
    # Stores the values into valueCatList
    for i in valueCatList.keys():
        valueCatList[i] = ENVResult[i]

    for key in valueCatList.keys():
        # Temp is used to store and modify the current value its on to remove percent signs
        temp = None
        try:
            temp = valueCatList[key].replace('%', '')

        # If the value is an integer then dont bother doing it
        except AttributeError:
            temp = valueCatList[key]

        # Make sure that all the values that can be numbers are numbers
        try:
            value = float(temp)

        # If it can't be a number then don't change anything
        except ValueError:
            value = valueCatList[key]

        # If the value is BDL then it has to be A
        if value == 'BDL*':
            results[key] = '0'

        # If the value is a string then find out what one it is and find the category by its type
        elif type(value) == str:
            # If the value of salmonella is negative then its A
            if value == 'Negative' and key == '17' or value == 'NEGATIVE' and key == '17':
                results[key] = '0'
            # If the value of salmonella is positive than it exceeds
            elif value == 'Positive' and key == '17' or value == 'POSITIVE' and key == '17':
                results[key] = '1'
            # If its Fecal Coliform and its <3 then its A
            elif value == '<3' and key == '16':
                results[key] = '0'
            # if its Fecal Coliform and its >1000 then it exceeds
            elif value == '>1000' and key == '16':
                results[key] = '1'

        elif value >= catMax[key]:
            results[key] = '1'
        else:
            results[key] = '0'
    # print results
    return results


def CCMEResults(CQARef):
    print Colors.bcolors.OKGREEN + (' EXECUTING CCME RESULTS') + Colors.bcolors.ENDC
    # ------------------------------------------------------Env Report------------------------------------------------------#
    # Dictonary for the environmental report values
    ENVDict = {
        "1": "Arsenic",  # 1-11, rep 1
        "2": "Cadmium",
        "3": 'Chromium',
        "4": 'Cobalt',
        "5": 'Copper',
        '6': 'Lead',
        "7": 'Mercury',
        '8': 'Molybdenum',
        '9': 'Nickel',
        '10': 'Selenium',
        '11': 'Zinc',
        '12': 'Total FM > 25 mm',
        '13': 'Total sharps > 2.8 mm*',
        '14': 'Total sharps > 12.5 mm',
        '15': 'Respiration-mgCO2-C/g OM/day',  # 15-17, report #1
        '16': 'Fecal Coliform',
        '17': 'Salmonella spp.',
        '19': 'Moisture',
        '18': 'Total Organic Matter',
        '20': 'Total Organic Matter @ 550 deg C',
        '22': 'C:N Ratio',
        '28': 'Total Solids (as received)',
        '30': 'Bulk Density (As Recieved)',
        '33': 'Ammonia (NH3/NH4-N)',
        '34': 'Total Phosphorus (As P205)',
        '35': 'Total Potassium (as K20)',
        '36': 'Calcium',
        '37': 'Magnesium',
        '38': 'Sulphur'
    }

    # connect to database
    cnx = SQL_CONNECTOR.test_connection()

    # Querry out the report numbers
    cursor = cnx.cursor()
    query = "SELECT * FROM alms.report WHERE refno='%s'" % (CQARef)
    cursor.execute(query)
    reportNumbers = []
    for item in cursor:
        print item
        tempList = [item[0], item[1]]
        reportNumbers.append(tempList)

    # Store the report numbers
    soilReport = None
    envReport = None
    for item in reportNumbers:
        if str(item[1]) == "SOIL":
            soilReport = str(item[0])
        elif str(item[1]) == "ENVI":
            print 'envi report found'
            envReport = str(item[0])

    print '==============================Here is the soil and env #================='
    print soilReport, envReport

    ENVResult = {}  # store all the required variables/parameters from the ENV report

    for key in ENVDict.keys():
        parameter = ENVDict[key]
        print key, parameter
        # query every parameter based on name
        if parameter != "Total FM > 25 mm":
            envQuery = r"""select
            ed.result_str
            from env_data as ed
            inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
            inner join report as r on ed.rptno=r.rptno
            inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
            left join units as u on ed.unit=u.un_units
            where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='%s'
            order by f.prt_sort,f.anly_sort""" % (envReport, parameter)

        # 12
        # query for specific unit of Total FM >25 mm parameter
        elif parameter == "Total FM > 25 mm":
            envQuery = r"""select
                ed.result_str
                from env_data as ed
                inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
                inner join report as r on ed.rptno=r.rptno
                inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
                left join units as u on ed.unit=u.un_units
                where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.unit = "pieces/500ml" and ed.rptno='%s' and f.rpt_dscrpt='%s'
                order by f.prt_sort,f.anly_sort""" % (envReport, parameter)

        cursor = cnx.cursor()
        cursor.execute(envQuery)
        # Store the querried information in the ENVResult list
        for item in cursor:
            print str(item[0])
            ENVResult[key] = str(item[0])

    print '----salmonella------'
    print ENVResult['17']
    print '=-----moisture------'
    print ENVResult['19']

    # Querry out moisture to calculate Results as recieved
    envQuery = r"""select
            ed.result_str
            from env_data as ed
            inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
            inner join report as r on ed.rptno=r.rptno
            inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
            left join units as u on ed.unit=u.un_units
            where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='Moisture'
            order by f.prt_sort,f.anly_sort""" % (envReport)

    cursor = cnx.cursor()
    cursor.execute(envQuery)
    # Store the querried information
    for item in cursor:
        moisture = item[0]

    for i in range(34, 39):
        temp = str(i)
        dm = 100 - float(moisture)
        dm = dm / 100
        ENVResult[temp] = str(float(ENVResult[temp]) * dm)

        # ------------------------------------------------------Sieves------------------------------------------------------#
    # 22
    # List of all possible Sieve sizes
    sieveList = [r'Sieve 2 Inch (% Passing)', r'Sieve 1 Inch (% Passing)', r'Sieve 1/2 Inch (% Passing)',
                 r'Sieve 3/8 Inch (% Passing)',
                 r'Sieve 1/4 Inch (% Passing)']

    sieveDict = {}
    sieveResult = {}
    sieveDifference = []
    targetPercent = 80.0
    smallestParameter = None
    # Querry out the sieve sizes
    for item in sieveList:
        parameter = item
        envQuery = r"""select
        ed.result_str
        from env_data as ed
        inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
        inner join report as r on ed.rptno=r.rptno
        inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
        left join units as u on ed.unit=u.un_units
        where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='%s'
        order by f.prt_sort,f.anly_sort""" % (envReport, parameter)
        cursor = cnx.cursor()
        cursor.execute(envQuery)
        # Goes through all sieve sizes
        for item in cursor:
            a = str(item[0])
            item = float(item[0])
            # Check to see if the current sieve size is below the target size and if it is then sets the value too 999
            if item - targetPercent <= 0:
                sieveDict[parameter] = 999
                sieveResult[parameter] = a
                # If its not below the target size then sets the value to the difference of the two numbers
            else:
                sieveDict[parameter] = item - targetPercent
                sieveResult[parameter] = a

    # get smallest size
    for key, value in sorted(sieveDict.iteritems(), key=lambda (k, v): (v, k)):  # get smallest size
        smallestParameter = key
        break

    temp1 = smallestParameter.replace('Sieve', '').replace('(% Passing)', '')[1:-1]

    # get sieve size
    ENVResult['22'] = temp1

    # ------------------------------------------------------Soil Report------------------------------------------------------#

    soilResult = {}

    # Querrys out the different soil report values
    # 20 and 29: PH
    soilResult['20'] = Utilities.getPH(CQARef)
    soilResult['29'] = Utilities.getPH(CQARef)

    # 23: salt
    query = """select
        IF(s.salt>0,s.salt,null) as salt
        from soil as s
        inner join report as r on r.rptno=s.rptno
        where s.rptno='%s'
        order by s.labno""" % (soilReport)

    cursor = cnx.cursor()
    cursor.execute(query)
    for item in cursor:
        soilResult['23'] = str(round(item[0], 1))

    print Colors.bcolors.OKCYAN + "----------------------------------CCME total organic-------------------------------" + Colors.bcolors.ENDC
    soilResult['32'] = Utilities.getNitrogen(CQARef)  # Nitrogen
    soilResult['18'] = Utilities.getTotalOrganicMatter(CQARef)  # Total organic MAtter
    available_matter_for_calc = Utilities.getAvailableOrganicMatter(CQARef)
    totalOrganicCarbon2 = organicCarbon(available_matter_for_calc)
    Nitrogen = float(soilResult['32'])
    print 'Total Organic            :' + str(soilResult['18'])
    print 'AVailable Organic        :' + str(available_matter_for_calc)
    print 'OG CARB * 0.6            :' + str(totalOrganicCarbon2)
    print 'Nitrogen                 :' + str(Nitrogen)

    # Divide organic carbon by nitrogen
    CNRatioValue = round((organicCarbon(available_matter_for_calc) / 0.9) / Nitrogen)
    print 'CNRatioValue             :' + str(CNRatioValue)

    cNRatio = str("%d:1" % (CNRatioValue))
    print 'Calculated CN Ratio      :' + cNRatio
    soilResult['21'] = cNRatio
    soilResult['31'] = cNRatio

    # ----------------------------Interpretation--------------------------------------------#

    # CEC Calculations (k_m3, mg_m3, ca_m3, na, buffer)
    # 24-27
    CECDict = {
        'k_m3': [],
        'mg_m3': [],
        'ca_m3': [],
        'na': [],
        'buffer': []
    }

    # Querry the cecDict
    for key in CECDict.keys():
        parameter = key
        query = """select
            IF(s.%s>0,s.%s,null) as %s
            from soil as s
            inner join report as r on r.rptno=s.rptno
            where s.rptno='%s'
            order by s.labno""" % (parameter, parameter, parameter, soilReport)

        cursor = cnx.cursor()
        cursor.execute(query)

        # If they don't exist then stop else add them to the dict
        for value in cursor:
            if value is None:
                sys.exit()
            else:
                CECDict[key].append(float(value[0]))

    # perk - k_m3, 390, E28
    # perMG - _mg_m3, 121.6, E29
    # perCa - ca_m3, 200.0, E30
    # perNa - na, 230.0., E26

    # Call the cakculateCEC function
    cec = calculateCEC(CECDict)

    # calculation parameters
    interpDict = {
        '24': ['na', 230.0],
        '25': ['k_m3', 390.0],
        '26': ['mg_m3', 121.6],
        '27': ['ca_m3', 200.0]
    }

    # add results to soilResults
    for key in interpDict.keys():
        # call def percentCalc(cec, paramterName, y, CECDict):
        result = percentCalc(cec, CECDict[interpDict[key][0]][0], interpDict[key][1], CECDict)
        soilResult[key] = result

        # ----------------------------Merging and Formatting--------------------------------------------#

    # Runs function that merges the two dict's
    tempResult = merge_two_dicts(ENVResult, soilResult)

    finalResult = {}
    # Goes through all the results
    for key in tempResult.keys():
        # Stores results into lists
        value = tempResult[key]
        digits = formatDict[key]

        # Sees what format the number needs and modifies the value to fit the format
        try:
            float(value)
            if digits == 2:
                finalValue = "%.2f" % (float(value))
                # print key, value, finalValue
                finalResult[key] = finalValue
            elif digits == 1:
                finalValue = "%.1f" % (float(value))
                # print key, value, finalValue
                finalResult[key] = finalValue
            elif digits == 0:
                finalValue = int(value)
                finalResult[key] = finalValue
            elif digits == '2%':
                finalValue = "%.2f" % (float(value)) + '%'
                finalResult[key] = finalValue

        # If there's a value error then don't change it
        except ValueError:
            finalResult[key] = value

    cursor.close()
    cnx.close()
    return finalResult
