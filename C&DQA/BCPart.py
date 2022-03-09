import os, sys, shutil
import mysql.connector
from mysql.connector import errorcode
from openpyxl.styles import Border, Side, PatternFill
from Utilities import FixFormatting, config, formatDict, organicCarbon, calculateCEC, percentCalc, merge_two_dicts


def BCPrint(Workbook, finalResult, CQARef):
    # Runs the method to check what values have failed standards
    highlightList = FailStandard(CQARef)
    # This is the sheet name inside of the template must be exact
    sheet = Workbook.get_sheet_by_name("CCME (British Columbia)")
    # Sets the color of the highlight/fill to highlight the faild values
    highlight = PatternFill(start_color='F3F315', end_color='F3F315', fill_type='solid')

    # Unicode to add the small 2's
    # sheet['A34'] = u'CO\u2082 Respiration Rate'
    # sheet['A36'] = u'O\u2082 Uptake Respiration Rate'

    # Prints the value to the excel sheet for each one by checking all of the boxes that should have a number in them
    print "----Trace Elements----"
    # D10-20: 1-11
    for i in range(10, 21, 1):
        # finds what value is currently in the cell(A1 for example)
        value = sheet.cell(row=i, column=3).value
        try:
            # puts the value of the cell into string format
            cellName = str(value)
            # uses the value to search for the result
            valueToFill = finalResult[cellName]
        except KeyError:
            # If there is'nt a result fill in a error
            valueToFill = "Error"

        # Fill in the valueToFill variable
        sheet.cell(row=i, column=3).value = valueToFill
        # Check if the value fails standards
        # if there is a none value for some reason skip it instead of crashing the program
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=3).fill = highlight

    print "----foreign matter----"
    # D25 - 27: 39-40 & 12
    for i in range(25, 27, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=4).value = valueToFill
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight

    print "----Maturity/Stability----"
    # D34: 15
    value = sheet.cell(row=30, column=4).value
    try:
        cellName = str(value)
        valueToFill = finalResult[cellName]
    except KeyError:
        valueToFill = "Error"

    sheet.cell(row=30, column=4).value = valueToFill
    if value == None:
        print ''
    elif highlightList[str(value)] == '1':
        sheet.cell(row=30, column=4).fill = highlight

    print"----Pathogens----"
    # D43-44: 16-17
    for i in range(37, 39, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=4).value = valueToFill
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight

    print"----CFIA----"
    # F54-55: 18-19
    for i in range(43, 45, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=4).value = valueToFill

    print"----Finished Compost Quality----"
    # f60-64: 20-24
    for i in range(51, 56, 1):
        value = sheet.cell(row=i, column=5).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=5).value = valueToFill

    # f66-68: 25-27
    for i in range(57, 60, 1):
        value = sheet.cell(row=i, column=5).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=5).value = valueToFill

    print"----Compost Agriculttral Product Value----"
    # D96-99: 28-31
    for i in range(95, 99, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=4).value = valueToFill

    # D101-107: 32-38
    for i in range(100, 107, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"

        sheet.cell(row=i, column=4).value = valueToFill

    # Below is all the code used to fix the borders by giving it a range of cell values(eg. A10:I10) and using Border to change the thickness size
    # Creates the border styles for the different side types
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    # Trace metals
    border = Border(bottom=thick, top=thick, left=thin, right=thin)
    FixFormatting(sheet, 'A6:B9', border)
    border = Border(bottom=thick, top=thick, right=thick)
    FixFormatting(sheet, 'C6:C9', border)
    border = Border(bottom=thick, top=thick, right=thin)
    FixFormatting(sheet, 'D6:D7', border)
    border = Border(bottom=thick, top=thick, right=thin)
    FixFormatting(sheet, 'E6:E7', border)
    border = Border(bottom=thick, top=thick, right=thin)
    FixFormatting(sheet, 'F6:G7', border)
    border = Border(bottom=thick, top=thick, right=thick)
    FixFormatting(sheet, 'H6:I7', border)
    border = Border(bottom=thick, right=thick)
    FixFormatting(sheet, 'D8:I9', border)
    border = Border(bottom=thick, right=thick)
    FixFormatting(sheet, 'A20:I20', border)

    border = Border(right=thick, bottom=thin)
    for i in range(10, 20):
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # Foreign Matter
    border = Border(bottom=thick, top=thick, right=thick)
    FixFormatting(sheet, 'A24:I24', border)
    border = Border(bottom=thin, right=thick)
    FixFormatting(sheet, 'A25:I25', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A26:I26', border)

    # Maturity and stability
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'A29:I29', border)
    border = Border(bottom=thin, right=thin)
    FixFormatting(sheet, 'A30:C31', border)
    border = Border(right=thin, bottom=thick)
    FixFormatting(sheet, 'A32:C33', border)
    border = Border(bottom=thin, right=thin)
    FixFormatting(sheet, 'D30:D31', border)
    border = Border(bottom=thick, right=thin)
    FixFormatting(sheet, 'D32:D33', border)
    border = Border(bottom=thin, right=thick)
    FixFormatting(sheet, 'E30:I31', border)
    border = Border(bottom=thick, right=thick)
    FixFormatting(sheet, 'E32:I33', border)

    # Pathogens
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'A36:I36', border)
    border = Border(right=thick, bottom=thin)
    FixFormatting(sheet, 'A37:I37', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A38:I38', border)

    # CFIA
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'A42:I42', border)
    border = Border(right=thick, bottom=thin)
    FixFormatting(sheet, 'A43:I43', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A44:I44', border)
    border = Border(right=thin)
    FixFormatting(sheet, 'A44:C44', border)
    border = Border(right=thin)
    FixFormatting(sheet, 'A43:C43', border)

    # Compost Quality
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'B50:H50', border)

    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'B59:H59', border)

    border = Border(right=thick, bottom=thin)
    for i in range(51, 59):
        cellNumber = 'B%s:H%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # reference Compost Quality
    border = Border(bottom=thick, top=thick)
    FixFormatting(sheet, 'A64:I64', border)
    border = Border(bottom=thick)
    FixFormatting(sheet, 'A74:I74', border)

    border = Border(bottom=thin)
    for i in range(65, 74):
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # Compost product value
    border = Border(bottom=thick, top=thick, right=thick)
    FixFormatting(sheet, 'A93:I93', border)
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A106:I106', border)

    border = Border(top=thin, right=thick, bottom=thin)
    for i in range(94, 106):
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # imports all the photos to the sheet
    from openpyxl.drawing.image import Image
    # The directory of where the photos are kept
    os.chdir(r'C:\CQA\NewFormatCQA\Photos')
    # AL
    # Creating an image variable
    img = Image('al.jpg')
    # Adding it to the sheet at A1
    sheet.add_image(img, 'A1')
    img2 = Image('al.jpg')
    sheet.add_image(img2, 'A46')
    img3 = Image('al.jpg')
    sheet.add_image(img3, 'A89')

    # Cp
    img = Image('cp.png')
    sheet.add_image(img, 'H1')
    img = Image('cp.png')
    sheet.add_image(img, 'H46')
    img = Image('cp.png')
    sheet.add_image(img, 'H89')

    saveLocation = os.path.join(r"C:\CQA\NewFormatCQA\FinishedReport", CQARef)

    Workbook.save(saveLocation + "\%sReport.xlsx" % (CQARef))


def findBCCatagory(CQARef):
    # BC is weird. going to have to check each class by compairing the results
    # And the catagory and see what one it falls under
    # gets the results from the Ontario Results Function
    ENVResult = BCResults(CQARef)
    # Puts the 17 results needed into a separate list called valueCatList
    valueCatList = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '13': '',
        '15': '',
        '16': '',
        '17': '',
        '18': '',
        '21': ''
    }

    catABio = {
        '1': 15,
        '2': 2,
        '3': -1,
        '4': 30,
        '5': -1,
        '6': 100,
        '7': 1,
        '8': 4,
        '9': 36,
        '10': 2.8,
        '11': 370,
        '39': 1,
        '13': 0,
        '15': 4,
        '16': 1000,
        '17': 3,
        '18': -1,
        '21': -1
    }
    catACom = {
        '1': 13,
        '2': 3,
        '3': 100,
        '4': 34,
        '5': 400,
        '6': 150,
        '7': 2,
        '8': 5,
        '9': 62,
        '10': 2,
        '11': 500,
        '39': 1,
        '13': 0,
        '15': 4,
        '16': 1000,
        '17': 3,
        '18': -1,
        '21': -1
    }

    catB = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '13': '',
        '15': '',
        '16': '',
        '17': '',
        '18': -1,
        '21': -1
    }
    # Declaring the final result
    finalResult = ''

    # Stores the values into valueCatList
    for i in valueCatList.keys():
        valueCatList[i] = ENVResult[i]

    for key in valueCatList.keys():
        # Temp is used to store and modify the current value its on to remove percent signs
        temp = None
        try:
            temp = valueCatList[key].replace('%', '')
        # If the value is an integer then dont bother doing it
        except AttributeError:
            temp = valueCatList[key]

        # Make sure that all the values that can be numbers are numbers
        try:
            value = float(temp)
        # If it can't be a number then don't change anything
        except ValueError:
            value = valueCatList[key]
    # Answer is used to check and break out of the loop if the value is over the
    # catagory number
    for key in valueCatList.keys():
        if catABio[key] == -1:
            continue
        elif value == 'BDL*':
            continue
        elif value <= catABio:
            continue
        else:
            break

    # goes through all the dictionary results and finds the highest category letter and returns that one
    for key in results.keys():
        parameter = results[key]

        if finalResult == parameter:
            finalResult = parameter
        elif finalResult == 'AA' and (parameter == 'A' or parameter == 'B' or parameter == 'Exceeds'):
            finalResult = parameter
        elif finalResult == 'A' and (parameter == 'B' or parameter == 'Exceeds'):
            finalResult = parameter
        elif finalResult == 'B' and parameter == 'Exceeds':
            finalResult = parameter

    return finalResult


def FailStandard(CQARef):
    # Classes/catagorys need to be changed due to changes with BC
    results = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '13': '',
        '15': '',
        '16': '',
        '17': '',
        '18': ''
    }
    # gets the results from the Quebec Results Function
    ENVResult = BCResults(CQARef)
    # Puts the 17 results needed into a separate list called valueCatList
    valueCatList = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '13': '',
        '15': '',
        '16': '',
        '17': '',
        '18': ''
    }
    # category max
    catMax = {
        '1': 75,
        '2': 20,
        '3': 1060,
        '4': 150,
        '5': 2200,
        '6': 500,
        '7': 15,
        '8': 20,
        '9': 180,
        '10': 14,
        '11': 1850,
        '39': 1,
        '13': 0,
        '15': 4,
        '16': 1000,
        '17': 3,
        '18': 15
    }
    # Stores the values into valueCatList
    for key in valueCatList.keys():
        valueCatList[key] = ENVResult[key]

    for key in valueCatList.keys():
        # Temp is used to store and modify the current value its on to remove percent signs
        temp = None
        try:
            temp = valueCatList[key].replace('%', '')
        # If the value is an integer then dont bother doing it
        except AttributeError:
            temp = valueCatList[key]

        # Make sure that all the values that can be numbers are numbers
        try:
            value = float(temp)
        # If it can't be a number then don't change anything
        except ValueError:
            value = valueCatList[key]

        # If the value is BDL then it has to be 0
        if value == 'BDL*':
            results[key] = '0'

        # If the value is a string then find out what one it is and find the category by its type
        elif type(value) == str:
            # If the value of salmonella is negative then its AA
            if value == 'Negative' and key == '17' or value == 'NEGATIVE' and key == '17':
                results[key] = '0'
            # If the value of salmonella is positive than it exceeds
            elif value == 'Positive' and key == '17' or value == 'POSITIVE' and key == '17':
                results[key] = '1'
            # If its E. Coli and its <3 then its AA
            elif value == '<3' and key == '16':
                results[key] = '0'
            # if its E. Coli and its >1000 then it exceeds
            elif value == '>1000' and key == '16':
                results = '1'

        elif value >= catMax[key]:
            results[key] = '1'
        else:
            results[key] = '0'
    return results


def BCResults(CQARef):
    # ------------------------------------------------------Env Report------------------------------------------------------#
    # Dictionary for the environmental report values
    ENVDict = {
        "1": "Arsenic",
        "2": "Cadmium",
        "3": 'Chromium',
        "4": 'Cobalt',
        "5": 'Copper',
        '6': 'Lead',
        "7": 'Mercury',
        '8': 'Molybdenum',
        '9': 'Nickel',
        '10': 'Selenium',
        '11': 'Zinc',
        '12': 'Total FM > 25 mm',
        '13': 'Total sharps > 2.8 mm*',
        '14': 'Total sharps > 12.5 mm',
        '15': 'Respiration-mgCO2-C/g OM/day',
        '16': 'E. coli',  # Change back to Fecal Coliform when working
        '17': 'Salmonella spp.',
        '19': 'Moisture',
        '20': 'Total Organic Matter @ 550 deg C',
        '22': 'C:N Ratio',
        '28': 'Total Solids (as received)',
        '30': 'Bulk Density (As Recieved)',
        '33': 'Ammonia (NH3/NH4-N)',
        '34': 'Total Phosphorus (As P205)',
        '35': 'Total Potassium (as K20)',
        '36': 'Calcium',
        '37': 'Magnesium',
        '38': 'Sulphur',
        '39': 'Total FM > 2.8 mm*',
        '40': 'Total plastics > 2.8 mm*'
    }

    # connect to database
    try:
        cnx = mysql.connector.connect(**config)
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)

    # Querry out the report numbers
    cursor = cnx.cursor()
    query = "SELECT * FROM alms.report WHERE refno='%s'" % (CQARef)
    cursor.execute(query)
    reportNumbers = []
    for item in cursor:
        print item
        tempList = [item[0], item[1]]
        reportNumbers.append(tempList)

    # Store the report numbers
    soilReport = None
    envReport = None
    for item in reportNumbers:
        if str(item[1]) == "SOIL":
            soilReport = str(item[0])
        elif str(item[1]) == "ENVI":
            envReport = str(item[0])

    print soilReport, envReport

    ENVResult = {}

    for key in ENVDict.keys():
        parameter = ENVDict[key]

        # query every parameter based on name
        if parameter != "Total FM > 25 mm":
            envQuery = r"""select
            ed.result_str
            from env_data as ed
            inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
            inner join report as r on ed.rptno=r.rptno
            inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
            left join units as u on ed.unit=u.un_units
            where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='%s'
            order by f.prt_sort,f.anly_sort""" % (envReport, parameter)

        # 12
        # query for specific unit of Total FM >25 mm parameter
        elif parameter == "Total FM > 25 mm":
            envQuery = r"""select
                ed.result_str
                from env_data as ed
                inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
                inner join report as r on ed.rptno=r.rptno
                inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
                left join units as u on ed.unit=u.un_units
                where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.unit = "pieces/500ml" and ed.rptno='%s' and f.rpt_dscrpt='%s'
                order by f.prt_sort,f.anly_sort""" % (envReport, parameter)

        cursor = cnx.cursor()
        cursor.execute(envQuery)
        # Store the querried information in the ENVResult list
        for item in cursor:
            ENVResult[key] = str(item[0])

    # Querry out moisture to calculate Results as recieved
    envQuery = r"""select
            ed.result_str
            from env_data as ed
            inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
            inner join report as r on ed.rptno=r.rptno
            inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
            left join units as u on ed.unit=u.un_units
            where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='Moisture'
            order by f.prt_sort,f.anly_sort""" % (envReport)

    cursor = cnx.cursor()
    cursor.execute(envQuery)
    # Store the querried information
    for item in cursor:
        moisture = item[0]

    for i in range(34, 39):
        temp = str(i)
        dm = 100 - float(moisture)
        dm = dm / 100
        ENVResult[temp] = str(float(ENVResult[temp]) * dm)
        # ------------------------------------------------------Sieves------------------------------------------------------#
    # 22
    # List of all possible Sieve sizes
    sieveList = [r'Sieve 2 Inch (% Passing)', r'Sieve 1 Inch (% Passing)', r'Sieve 1/2 Inch (% Passing)',
                 r'Sieve 3/8 Inch (% Passing)',
                 r'Sieve 1/4 Inch (% Passing)']

    sieveDict = {}
    sieveResult = {}
    targetPercent = 80.0
    smallestParameter = None
    # Querry out the sieve sizes
    for item in sieveList:
        parameter = item
        envQuery = r"""select
        ed.result_str
        from env_data as ed
        inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
        inner join report as r on ed.rptno=r.rptno
        inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
        left join units as u on ed.unit=u.un_units
        where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='%s'
        order by f.prt_sort,f.anly_sort""" % (envReport, parameter)
        cursor = cnx.cursor()
        cursor.execute(envQuery)
        # Goes through all sieve sizes
        for item in cursor:
            a = str(item[0])
            item = float(item[0])
            # Check to see if the current sieve size is below the target size and if it is then sets the value too 999
            if item - targetPercent <= 0:
                sieveDict[parameter] = 999
                sieveResult[parameter] = a
                # If its not below the target size then sets the value to the difference of the two numbers
            else:
                sieveDict[parameter] = item - targetPercent
                sieveResult[parameter] = a

    # get smallest size
    for key, value in sorted(sieveDict.iteritems(), key=lambda (k, v): (v, k)):
        smallestParameter = key
        break

    tempSieve = smallestParameter.replace('Sieve', '').replace('(% Passing)', '')[1:-1]

    # sets the sieve size
    ENVResult['22'] = tempSieve

    # ------------------------------------------------------Soil Report------------------------------------------------------#

    soilResult = {}

    # Querrys out the different soil report values
    # 20 and 29: PH
    query = """select
            IF(s.ph>0,s.ph,null) as ph
            from soil as s
            inner join report as r on r.rptno=s.rptno
            where s.rptno='%s'
            order by s.labno""" % (soilReport)

    cursor.execute(query)
    for item in cursor:
        soilResult['20'] = str(item[0])
        soilResult['29'] = str(item[0])

    # 23: salt
    query = """select
        IF(s.salt>0,s.salt,null) as salt
        from soil as s
        inner join report as r on r.rptno=s.rptno
        where s.rptno='%s'
        order by s.labno""" % (soilReport)

    cursor = cnx.cursor()
    cursor.execute(query)
    for item in cursor:
        soilResult['23'] = str(round(item[0], 1))

        # 32 Nitrogen(total)
    query = """select
        IF(s.result>0,s.result,null) as result
        from agdata as s
        inner join report as r on r.rptno=s.rptno
        where s.rptno='%s'
        order by s.labno""" % (soilReport)

    cursor = cnx.cursor()
    cursor.execute(query)
    for item in cursor:
        soilResult['32'] = str(item[0])

    # -----------Total Organic matter-------------#
    query = """select
        IF(s.om>0,s.om,null) as om
        from soil as s
        inner join report as r on r.rptno=s.rptno
        where s.rptno='%s'
        order by s.labno""" % (soilReport)

    cursor = cnx.cursor()
    cursor.execute(query)

    for item in cursor:
        # Total Organic Matter = Organic Matter / 0.9
        temp = float(item[0])
        soilResult['18'] = str(temp / 0.9)

    # 21 & 31 C:N ratio
    # Round totalOrganicCarbin and Nitrogen to be a whole number
    totalOrganicCarbon = round(organicCarbon(soilResult['18']))
    Nitrogen = round(float(soilResult['32']))

    # Divide organic carbon by nitrogen
    if Nitrogen > 1:
        totalOrganicCarbon = round(organicCarbon(soilResult['18']) / float(soilResult['32']))

    cNRatio = str("%d:1" % (totalOrganicCarbon))

    soilResult['21'] = cNRatio
    soilResult['31'] = cNRatio

    # ----------------------------Interpretation--------------------------------------------#

    # CEC Calculations (k_m3, mg_m3, ca_m3, na, buffer)
    # 24-27
    CECDict = {
        'k_m3': [],
        'mg_m3': [],
        'ca_m3': [],
        'na': [],
        'buffer': []
    }
    # Querry the cecDict
    for key in CECDict.keys():
        parameter = key
        query = """select
            IF(s.%s>0,s.%s,null) as %s
            from soil as s
            inner join report as r on r.rptno=s.rptno
            where s.rptno='%s'
            order by s.labno""" % (parameter, parameter, parameter, soilReport)

        cursor = cnx.cursor()
        cursor.execute(query)
        # If they don't exist then stop else add them to the dict
        for value in cursor:
            if value is None:
                sys.exit()
            else:
                CECDict[key].append(float(value[0]))

    # perk - k_m3, 390, E28
    # perMG - _mg_m3, 121.6, E29
    # perCa - ca_m3, 200.0, E30
    # perNa - na, 230.0., E26

    # Call the cakculateCEC function
    cec = calculateCEC(CECDict)

    # calculation parameters
    interpDict = {
        '24': ['na', 230.0],
        '25': ['k_m3', 390.0],
        '26': ['mg_m3', 121.6],
        '27': ['ca_m3', 200.0]
    }

    # add results to soilResults
    for key in interpDict.keys():
        # call def percentCalc(cec, paramterName, y, CECDict):
        result = percentCalc(cec, CECDict[interpDict[key][0]][0], interpDict[key][1], CECDict)
        soilResult[key] = result

        # ----------------------------Merging and Formatting--------------------------------------------#

    # Runs function that merges the two dict's
    tempResult = merge_two_dicts(ENVResult, soilResult)

    finalResult = {}
    # Goes through all the results
    for key in tempResult.keys():
        # Stores results into lists
        value = tempResult[key]
        digits = formatDict[key]

        # Sees what format the number needs and modifies the value to fit the format
        try:
            float(value)
            if digits == 2:
                finalValue = "%.2f" % (float(value))
                # print key, value, finalValue
                finalResult[key] = finalValue
            elif digits == 1:
                finalValue = "%.1f" % (float(value))
                # print key, value, finalValue
                finalResult[key] = finalValue
            elif digits == 0:
                finalValue = int(value)
                finalResult[key] = finalValue
            elif digits == '2%':
                finalValue = "%.2f" % (float(value)) + '%'
                finalResult[key] = finalValue

        # If there's a value error then don't change it
        except ValueError:
            finalResult[key] = value

    cursor.close()
    cnx.close()
    return finalResult
