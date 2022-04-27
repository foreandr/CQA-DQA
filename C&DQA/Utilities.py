import collections
import os
import shutil
from collections import Counter
from datetime import time

import openpyxl
from openpyxl.styles import Border
import mysql.connector
from mysql.connector import errorcode
import SQL_CONNECTOR

import Colors

config = {
    'user': 'lmsuser',
    'password': 'readonly',
    'host': '10.0.0.26',
    'database': 'alms',
    'buffered': True
}
# Format dictonary for the final results number formatting (2=0.00, 1=0.0, 0=0, 'str'=A String, and '2%'=0.00%
formatDict = {
    "1": 2,
    "2": 2,
    "3": 2,
    "4": 2,
    "5": 2,
    "6": 2,
    '7': 2,
    '8': 2,
    '9': 2,
    '10': 2,
    '11': 2,
    '12': 0,
    '13': 0,
    '14': 0,
    '15': 2,
    '16': 0,
    '17': 'str',
    '18': '2%',
    '19': '2%',
    '20': 1,
    '21': 'str',
    '22': 'str',
    '23': 1,
    '24': '2%',
    '25': '2%',
    '26': '2%',
    '27': '2%',
    '28': '2%',
    '29': 0,
    '30': 0,
    '31': 'str',
    '32': '2%',
    '33': 2,
    '34': '2%',
    '35': '2%',
    '36': '2%',
    '37': '2%',
    '38': 2,
    '39': '2%',
    '40': '2%'
}
import math


def rpt_name_refno():
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = '''
    SELECT report.rpt_name,
        report.custno,
        report.module,
        report.rptno,
        report.company,
        report.grow_1,
        report.refno,
        report.rpt_status,
        report.create_date,
        report.state
        FROM alms.report report
        WHERE (report.rpt_name = 'SQA_COMP'
        OR report.rpt_name='AL_CQA-O'
        OR report.rpt_name = "AL-ON-CQ"
        OR report.rpt_name = "AL_CQA"
        OR report.rpt_name = "A&L-WD")
        AND rpt_status ="6" or rpt_status ="5"
    '''
    cursor.execute(query)
    dqaArray = []
    cqaArray = []
    for item in cursor:
        if item[0] == 'A&L-WD' and item[6] != '':
            dqaArray.append(item)
        else:
            if item[6] != '':
                cqaArray.append(item)
    print('\nCQA ARRAY')
    for i in cqaArray:
        print(i)

    print('\nDQA ARRAY')
    for i in dqaArray:
        print(i)


def miniInterpreter(string):
    new_string = string[1:]
    get_cell = new_string[:3]
    get_operator = new_string[3]

    string_list = string.split('%s' % get_operator)

    # print(string_list)
    amount = string_list[1]
    ##print(new_string)
    # print(get_cell)
    # rint(get_operator)

    # print(amount)
    return get_cell, amount, get_operator


def get_names_and_indexes(sheet):
    list_indexes_names = []
    for row in range(1, 100):
        if sheet.cell(row=row, column=1).value != None:
            text_value = sheet.cell(row=row, column=1).value
            # print row, text_value, 1
            temp_list = ['INDEX:', row, text_value]
            list_indexes_names.append(temp_list)
        else:
            text_value = sheet.cell(row=row, column=2).value
            # print row, text_value, 2
            temp_list = ['INDEX:', row, text_value]
            list_indexes_names.append(temp_list)
    newlist = []
    for i in range(len(list_indexes_names)):
        if list_indexes_names[i][1] != None:
            newlist.append(list_indexes_names[i])

    # print '\nprinting updated list\n'

    for i in newlist:
        print i

    return newlist


def add_round_to_excel_formula(string):
    if string == None or string == '':
        return string
    string_no_equals = string[1:]
    new_string = 'ROUND(' + string_no_equals + ', 2)'
    final_string_with_equals = '=' + new_string
    return (final_string_with_equals)


def grab_excel_locations():
    from openpyxl.styles import PatternFill, Border, Side
    from openpyxl import Workbook
    import openpyxl

    templateFile = r'C:\CQA\FULL CQA - DQA\C&DQA\Templates\Ontario DQA - KO.xlsx'
    wb = openpyxl.load_workbook(templateFile)
    sheet = wb.get_sheet_by_name("Ontario CFIA", )

    value_list = []
    for i in range(1, 90):
        column1_value = sheet.cell(row=i, column=1).value
        column2_value = sheet.cell(row=i, column=2).value

        if column1_value == '' or column1_value == None:
            value_list.append([i, column2_value])
        else:
            value_list.append([i, column1_value])
    final_list = []
    for i in value_list:
        if i[1] == None:
            continue
        else:
            final_list.append(i)

    for i in final_list:
        print(i)

    return final_list


def organic_matter_query(CQAREF):
    # print(CQAREF)
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = '''
    select om, refno
    from soil s
    INNER JOIN report rep
    ON rep.rptno = s.rptno
    where refno = '%s'
    ''' % CQAREF
    cursor.execute(query)
    value = 0
    for item in cursor:
        value = float(item[0])
        print(item)
    return value


def open_report_csv_INDEXLIST():
    import csv

    list_indexes_names = []
    with open('C:\CQA\FULL CQA - DQA/report_template.csv') as file:
        my_reader = csv.reader(file, delimiter=',')
        index = 0
        for i in my_reader:
            if i[1] == 'Trace Elements':
                # print i[1], ' GOT TRACE ELEMENTS'
                index = 4  # increasing to fix numbers
            if i[1] != '':
                templist = [index, i[1], 0]
                list_indexes_names.append(templist)
                # print index, i[1]
            if i[0] != '':
                templist = [index, i[0], 1]
                list_indexes_names.append(templist)
                # print index, i[0]
            index += 1

    # for i in list_indexes_names:
    #    print i
    return list_indexes_names


def associate_nums_with_values(valueDict, nameDict):
    newDict = {}
    for i in nameDict:
        if i in valueDict:
            title = nameDict[i]
            value = valueDict[i]
            temp_dict = {title: value}
            newDict.update(temp_dict)
    return newDict


def get_until_space(string):
    curr_string = ''
    for i in string:
        if i != ' ':
            curr_string += i
        else:
            return curr_string
    return curr_string


def makeDirectory(saveLocation_):
    try:
        os.mkdir(saveLocation_)
    except:
        shutil.rmtree(saveLocation_)
        os.mkdir(saveLocation_)


def return_worst(dict):
    stored_worst = ''
    for i in dict:
        if dict[i] == 'Fail':
            return 'Fail'
        elif dict[i] == 'B':
            stored_worst = 'B'
        elif dict[i] == 'A' and stored_worst != 'B':
            stored_worst = 'A'
        elif dict[i] == 'AA' and stored_worst != 'B' and stored_worst != 'A':
            stored_worst = 'AA'
    return stored_worst


def merge_two_dicts(x, y):
    '''Given two dicts, merge them into a new dict as a shallow copy.'''

    z = x.copy()
    z.update(y)
    return z


def calculateCEC(CECDict):
    '''take in a dictionary of 5 values (k_m3, mg_m3, ca_m3, na, buffer)'''

    k_m3 = CECDict['k_m3'][0]
    mg_m3 = CECDict['mg_m3'][0]
    ca_m3 = CECDict['ca_m3'][0]
    na = CECDict['na'][0]
    buffer = CECDict['buffer'][0]

    temp = k_m3 / 390.0 + mg_m3 / 121.6 + ca_m3 / 200.0 + na / 230.0

    if 6.6 > buffer and temp > 0:
        cec = temp + 4 * (6.6 - buffer)
        return cec
    else:
        return temp


def percentCalc(cec, paramterName, y, CECDict):
    '''takes the cec, the name of parameter, calculation of the parameter, and the dictionary'''

    k_m3 = CECDict['k_m3'][0]
    mg_m3 = CECDict['mg_m3'][0]
    ca_m3 = CECDict['ca_m3'][0]
    na = CECDict['na'][0]
    buffer = CECDict['buffer'][0]

    if cec > 0:
        result = float(paramterName) / float(y) / float(cec) * 100
        return result
    else:
        return None


def organicCarbon(value):
    '''Takes in the total organic matter and times it by 0.6 to make organic carbon'''

    result = float(value) * 0.6
    return result


def getTotalOrganicMatter(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """select env.result_str
        from alms.env_data env
        inner join alms.report rep
        on env.rptno = rep.rptno
        where env.feecode = 'GOMZ405' and rep.refno = '%s'""" % CQAREF
    cursor.execute(query)
    for item in cursor:
        # print('printing temp', item)
        if item == None:
            return 'ERROR'
        else:
            temp = item[0]
        # print temp
    return temp


def getAvailableOrganicMatter(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """select s.om 
    from soil s 
    inner join report r 
    on r.rptno=s.rptno 
    where r.refno = '%s'""" % CQAREF
    cursor.execute(query)
    temp = 0
    for item in cursor:
        temp = float(item[0])
        # print temp
    return temp


def removePercentSign(string):
    try:
        if string[-1] == '%':
            new_string = string[:-1]
            return new_string
    except:
        return string


def getValuesForAGIndex(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = '''
    select na, k, ph,cl, ca
    from agdata a
    inner join report r
    on r.rptno = a.rptno
    inner join soil s
    on s.rptno = a.rptno
    where route_4 = '%s'
    ''' % CQAREF
    cursor.execute(query)

    value_list = []
    for item in cursor:  # some weird data type gets pulled here
        value_list.append(item[0])
        value_list.append(item[1])
        value_list.append(item[2])
        value_list.append(item[3])
        value_list.append(item[4])

    item_dict = {}
    for i in range(len(value_list)):
        if i == 0:
            item_dict['NA'] = value_list[i]
        if i == 1:
            item_dict['K'] = value_list[i]
        if i == 2:
            item_dict['PH'] = value_list[i]
        if i == 3:
            item_dict['CL'] = value_list[i]
        if i == 4:
            item_dict['CA'] = value_list[i]
    return item_dict


def getNitrogen(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """select result 
    from agdata a 
    inner join report r 
    on r.rptno = a.rptno
    where route_4 = '%s'""" % CQAREF
    cursor.execute(query)
    for item in cursor:
        nitrogen = str(item[0])
    return nitrogen


def round_all_array_values(array):
    for row in array:
        try:
            row[2] = float(row[2])
            row[2] = round(row[2], 2)
        except:
            # print('cant cast to float')
            row[2] = row[2]


def getCO2Resp(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """
    SELECT result_str
    FROM env_data env
    inner join report rep
    on rep.rptno = env.rptno
    where feecode = 'GGCC642' and rep.refno = '%s'
    """ % CQAREF
    cursor.execute(query)
    for item in cursor:
        CO2 = str(item[0])
    return CO2


def getPH(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """
            select ph
            from soil s
            inner join report r on r.rptno=s.rptno
            where r.refno='%s'""" % (CQAREF)
    cursor.execute(query)
    for i in cursor:
        # print 'current ph' + str(i)
        ph = i[0]
    return ph


def getCNRatio(CQAREF):
    pass


def FixFormatting(ws, cell_range, border=Border()):
    '''Takes the worksheet, the cell range and the border type to fix border formatting issues'''

    # Set all the border types equal to themselves
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    # the cell range is the rows
    rows = ws[cell_range]

    # Goes through each cell in the range and sets the border for top, bottom, left, and right
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right


def findLocation(CQARef):
    '''find location to sort the different reports'''

    # connects to sql database
    cnx = SQL_CONNECTOR.test_connection()

    # Querrys the location none is a place holder, initally there is no value]
    # after the alms query is complete the location value is placed there
    location = None
    query = """select r.state
        from alms.report as r
        where r.refno='%s'""" % (CQARef)
    cursor = cnx.cursor()
    cursor.execute(query)
    # Sets the location and returns it as letters
    for item in cursor:
        location = str(item[0])

    cursor.close()
    cnx.close()
    return location


def full_location(CQAref):
    location = findLocation(CQAref)
    if location == 'ON':
        return 'Ontario'
    elif location == 'NB':
        return 'New Brunswick'
    elif location == 'QC' or location == 'QB':
        return 'Quebec'
    elif location == 'BC':
        return 'British Columbia'
    elif location == 'AB':
        return 'Alberta'
    elif location == 'SK':
        return 'Saskatchewan'
    elif location == 'NS':
        return 'Nova Scotia'
    elif location == 'PE':
        return 'Prince Edward Island'
    else:
        return location


def get_reference_numbers():
    connection = SQL_CONNECTOR.test_connection()
    print Colors.bcolors.UNDERLINE + Colors.bcolors.HEADER + "Executing Kelly's Queries" + Colors.bcolors.ENDC
    reference_number_index = 6
    cursor = connection.cursor()
    query = '''
    SELECT report.rpt_name, report.custno, report.module, report.rptno, report.company, report.grow_1, report.refno, report.rpt_status, report.create_date, report.state
    FROM alms.report report
    WHERE (report.rpt_name = 'SQA_COMP'
    OR report.rpt_name ='STP'
    OR report.rpt_name = 'AL_STP'
    OR report.rpt_name='AL_CQA-O'
    OR report.rpt_name = "AL-ON-CQ"
    OR report.rpt_name = "AL_CQA")
    AND ((rpt_status ="5" OR rpt_status ="6") AND refno !=  "")
    ORDER BY report.refno
    
    '''
    cursor.execute(query)
    returned_values = []
    for i in cursor:
        # print(i)
        returned_values.append(i[6])

    final_dict = Counter(returned_values)
    print(Colors.bcolors.OKGREEN + str(final_dict) + Colors.bcolors.ENDC)

    final_array = []
    for key, value in final_dict.items():
        if value == 2:
            final_array.append(key)

    REALLY_FINAL = []
    for i in final_array:
        if i.startswith('STP'):
            pass
        else:
            REALLY_FINAL.append(i)

    print('Printing the 2 arrays')
    print('OG   ' + str(final_array) + '\n')
    print('FINAL' + str(REALLY_FINAL) + '\n')

    return REALLY_FINAL


def get_fecal(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """
    SELECT result_str
    FROM env_data env
    inner join report rep
    on rep.rptno = env.rptno
    where feecode = 'BMPZ484' and rep.refno = '%s'
    """ % CQAREF
    cursor.execute(query)
    for item in cursor:
        fecal = str(item[0])
    return fecal


CQA_ON_DATA_CATEGORY = [
    ['Arsenic', 13, 13, 75],
    ['Cadmium', 3, 3, 20],
    ['Chromium', 210, 210, 1060],
    ['Cobalt', 34, 34, 150],
    ['Copper', 100, 400, 760],
    ['Lead', 150, 150, 500],
    ['Mercury', 0.8, 0.8, 5],
    ['Molybdenum', 5, 5, 20],
    ['Nickel', 62, 62, 180],
    ['Selenium', 2, 2, 14],
    ['Zinc', 500, 700, 1850],
]
CQA_NON_ON_DATA_CATEGORY = [
    ['Arsenic', 13, 75],
    ['Cadmium', 3, 20],
    ['Chromium', 210, 210],  # AMBIGUOUS
    ['Cobalt', 34, 150],
    ['Copper', 400, 400],  # AMBIGUOUS
    ['Lead', 150, 500],
    ['Mercury', 0.8, 5],
    ['Molybdenum', 5, 20],
    ['Nickel', 62, 180],
    ['Selenium', 2, 14],
    ['Zinc', 700, 1850],
    ['Fecal Coliform (MPN/g dry)', 1000]
]
# IF YOU EDIT THIS OR THE TITLES IN THE EXCELS OTHER PARTS HAVE TO CHANGE TOO
CQA_ON_SECOND_PART_CHECK = [
    [0, 'Total FM > 2.8 mm*', 1, 2],  # DONE
    [1, 'Total plastics > 2.8 mm*', 0.5, 0.5],  #
    [2, 'Total FM > 25 mm', 0.0, 0.0],
    [3, 'Total sharps > 2.8 mm*', 0.0, 3],
    [4, 'Total sharps > 12.5 mm', 0.0, 0],
    [5, 'Respiration-mgCO2-C/g OM/day', 0, 4],  # ABOVE 4 IS FAIL
    [6, 'E. coli', 0, 1000],  # ANYTHING LOWER 1000 'fix for symbols'
    [7, 'Salmonella spp.', 0, 1000],  # IF NOT NEG FAIL
]


def get_company_name(CQAref):
    # Finds the company name
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT company FROM alms.report WHERE refno='%s' """ % CQAref
    cursor.execute(query)
    company = ''
    for item in cursor:
        company = str(item[0])
    return company


def get_company_address(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT address1 FROM alms.report WHERE refno='%s' """ % CQAref
    cursor.execute(query)
    address = ''
    for item in cursor:
        address = str(item[0])
    return address


def get_company_city(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT city FROM alms.report WHERE refno='%s' """ % CQAref
    cursor.execute(query)
    city = ''
    for item in cursor:
        city = str(item[0])
    return city


def get_company_state(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT state FROM alms.report WHERE refno='%s'""" % CQAref
    cursor.execute(query)
    state = ''
    for item in cursor:
        state = str(item[0])
    return state


def BDL_PERCENT_check(value):
    if value == 'BDL' or value == 'bdl':
        return 'BDL'
    elif value == 'N/A' or value == 'n/a':
        return 'N/A'
    else:
        return "{:.2f}".format(float(value))


def get_company_state_v2(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT name FROM (alms.report INNER JOIN alms.provinces ON report.state = provinces.abbreviation) WHERE refno='%s' """ % CQAref
    cursor.execute(query)
    state = ''
    for item in cursor:
        # print Colors.bcolors.OKBLUE + "Province: " + str(item)
        state = item[0].encode('utf-8').strip()  # VERY IMPORTANT
    return state


def get_company_zipcode(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT zip FROM alms.report WHERE refno='%s' """ % CQAref
    cursor.execute(query)
    zipC = ''
    for item in cursor:
        zipC = str(item[0])
    return zipC


def get_FULL_ADDRESS(CQAref):
    zip = get_company_zipcode(CQAref)
    state = get_company_state(CQAref)
    city = get_company_city(CQAref)

    address = city + ', ' + state + ' ' + zip
    return address


def remove_greater_than(string):
    if string[0] == '>':
        try:
            return float(string[1:])
        except:
            return string[1:]
    else:
        return string


def get_sample_ID(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """SELECT grow_1 FROM alms.report WHERE refno='%s' """ % CQAref
    cursor.execute(query)
    for item in cursor:
        sampleID = str(item[0])
    return sampleID


def get_current_Date(CQAref):
    currentDate = time.localtime()[0:3]
    return currentDate


def get_feecode(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """
        SELECT report.refno, feedstock.description
        FROM (alms.env_samp env_samp
        INNER JOIN alms.feedstock feedstock
        ON (env_samp.feedstock_code = feedstock.code))
        INNER JOIN alms.report report ON (env_samp.rptno = report.rptno)
        WHERE (report.refno = '%s')""" % CQAref

    wanted_description = "Not Specified"
    cursor.execute(query)
    for item in cursor:
        wanted_description = item[1]  # second item in tuple is description, first is iD
    return wanted_description


def get_rptno(CQAref):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """
        select rep.rptno
        from env_data env
        inner join report rep
        on rep.rptno = env.rptno
        where rep.refno = '%s'
        """ % CQAref
    cursor.execute(query)
    for item in cursor:
        RPTNO = item[0]
    return RPTNO


def number_formatting(sheet):
    print(Colors.bcolors.OKBLUE + 'EXECUTION DQA PART A' + Colors.bcolors.ENDC)
    for i in range(10, 21):
        test_value = sheet.cell(row=i, column=4).value
        # print(test_value, type(test_value))
        if test_value == 'BDL':
            # print(test_value, 'GOT IT')
            sheet.cell(row=i, column=10).value = 'N/A'

    print(Colors.bcolors.OKBLUE + '\nPRINTING DQA EXECUTION' + Colors.bcolors.ENDC)
    array_of_values = []

    # print(sheet.cell(row=78, column=6).value)

    # GET VALUES IN A 2D ARRAY FOR CHECKING PURPOSES
    for i in range(71, 91):
        current_value = removePercentSign(sheet.cell(row=i, column=4).value)
        if current_value == None:
            new_value = float(sheet.cell(row=i, column=4).value)
            array_of_values.append([i, new_value])
        else:
            array_of_values.append([i, current_value])
        # print(removePercentSign(sheet.cell(row=i, column=4).value), sheet.cell(row=i, column=4).value)

    for i in array_of_values:
        # print(i)
        if i[1] == 0.0 or i[1] == '0.0' or i[1] < 0.001:
            # print('YES')
            sheet.cell(row=i[0], column=6).value = '0.0'
            sheet.cell(row=i[0], column=8).value = '0.0'
            sheet.cell(row=i[0], column=9).value = '0.0'
            sheet.cell(row=i[0], column=11).value = '0.0'

    AVAILABLE_S = sheet.cell(row=82, column=4).value
    sheet.cell(row=82, column=6).value = AVAILABLE_S / 0.1
    sheet.cell(row=82, column=8).value = AVAILABLE_S * 20
    sheet.cell(row=82, column=9).value = AVAILABLE_S * 0.01
    sheet.cell(row=82, column=11).value = AVAILABLE_S * 100


def get_cec_values(CQAREF):
    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = '''
    select s.k_m3, s.mg_m3, s.ca_m3, s.na, s.buffer
    from agdata a
    inner join report r
    on r.rptno = a.rptno
    inner join soil s
    on s.rptno = a.rptno
    where route_4 = '%s' 
    ''' % CQAREF
    cursor.execute(query)

    list_of_items = []
    for row in cursor:
        for item in row:
            list_of_items.append(float(item))

    return list_of_items


def andres_cec_calc(cec_array):
    value = (cec_array[0] / 390) + (cec_array[1] / 121.6) + (cec_array[2] / 200) + (cec_array[3] / 230.0)
    buffer = cec_array[4]

    if 6.6 > buffer and value > 0:
        cec = value + 4 * (6.6 - buffer)
        return cec
    else:
        return value


def andres_percent_calc(cec_value, array):
    k_m3_value = round((array[0] / 390 / cec_value * 100), 2)
    mg_m3_value = round((array[1] / 121.6 / cec_value * 100), 2)
    ca_m3_value = round((array[2] / 200 / cec_value * 100), 2)
    na = round((array[3] / 230 / cec_value * 100), 2)
    return (k_m3_value, mg_m3_value, ca_m3_value, na)

# print(andres_percent_calc(andres_cec_calc(get_cec_values('CQA2200119')), get_cec_values('CQA2200119')))
