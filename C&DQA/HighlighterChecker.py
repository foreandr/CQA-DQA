import openpyxl
from openpyxl.styles import PatternFill
import CQAUtilities
import Utilities


def try_cast_to_num(value):
    if value == '' or len(str(value)) == 0:
        return ''
    try:
        float(value)
        return float(value)
    except:
        return value


def get_ontario_cqa_constraints_A(sheet):
    highlight = PatternFill(start_color='F3F315', end_color='F3F315', fill_type='solid')
    # sheet.cell(row=9, column=4).value = 14 TESTER
    for i in range(9, 20):
        column4_value = try_cast_to_num(sheet.cell(row=i, column=4).value)
        column6_value = sheet.cell(row=i, column=6).value
        if type(column4_value) == unicode:
            continue
        if column4_value > column6_value:
            sheet.cell(row=i, column=4).fill = highlight
            # print('CURRENT TYPE', type(column4_value), column4_value)

    column4_value = try_cast_to_num(sheet.cell(row=24, column=4).value)
    column6_value = 1  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass

    elif column4_value > column6_value:
        sheet.cell(row=24, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=25, column=4).value)
    column6_value = 0.5  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=25, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=26, column=4).value)
    column6_value = 0  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=26, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=28, column=4).value)
    column6_value = 0  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=28, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=29, column=4).value)
    column6_value = 0  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=29, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=33, column=4).value)
    column6_value = 4  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif float(column4_value) > column6_value:
        sheet.cell(row=33, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=35, column=4).value)
    column6_value = 400  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value >= column6_value:
        sheet.cell(row=35, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=40, column=4).value)
    column6_value = 1000  # COMPARISON VALUE
    if column4_value == 'BDL' or column4_value == '<3':
        pass
    elif column4_value >= column6_value:
        sheet.cell(row=40, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=41, column=4).value)
    column6_value = 3  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    try:
        casted_value = float(column4_value)
    except:
        casted_value = column4_value
    if type(casted_value) == float:
        if casted_value >= column6_value:
            sheet.cell(row=41, column=4).fill = highlight
    else:
        print'NOT A NUMBER', (casted_value)

    if sheet.cell(row=55, column=6).value == 'N/A':
        sheet.cell(row=55, column=6).fill = highlight


def get_non_ontario_cqa_constraints(sheet):
    highlight = PatternFill(start_color='F3F315', end_color='F3F315', fill_type='solid')
    for i in range(10, 21):
        column4_value = try_cast_to_num(sheet.cell(row=i, column=4).value)
        column5_value = sheet.cell(row=i, column=5).value
        if type(column4_value) == unicode:
            continue
        if column4_value > column5_value:
            sheet.cell(row=i, column=4).fill = highlight
            # print('CURRENT TYPE', type(column4_value), column4_value)

    column4_value = try_cast_to_num(sheet.cell(row=26, column=4).value)
    column5_value = 1  # COMPARISON VALUE
    column6_value = 2
    if column4_value == 'BDL':
        pass
    elif column4_value > column5_value:
        sheet.cell(row=26, column=4).fill = highlight
    # elif column4_value > column6_value:

    column4_value = try_cast_to_num(sheet.cell(row=29, column=4).value)
    column6_value = 0  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=29, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=30, column=4).value)
    column6_value = 3  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=30, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=34, column=4).value)
    column6_value = 4  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    elif float(column4_value) > column6_value:
        sheet.cell(row=34, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=36, column=4).value)
    column6_value = 400  # COMPARISON VALUE
    if column4_value == 'BDL' or column4_value == '':
        pass
    elif column4_value > column6_value:
        sheet.cell(row=36, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=41, column=4).value)
    column6_value = 1000  # COMPARISON VALUE
    try:
        if column4_value == 'BDL' or column4_value == '<3':
            pass
        elif float(column4_value) >= float(column6_value):
            sheet.cell(row=41, column=4).fill = highlight
    except:
        print('got to here')
        if float((column4_value[1:])) >= float(column6_value):
            sheet.cell(row=41, column=4).fill = highlight

    column4_value = try_cast_to_num(sheet.cell(row=42, column=4).value)
    column6_value = 3  # COMPARISON VALUE
    if column4_value == 'BDL':
        pass
    try:
        casted_value = float(column4_value)
    except:
        casted_value = column4_value
    if casted_value == 'POSITIVE' or casted_value == 'positive':
        CAT_B = True
    if type(casted_value) == float:
        if casted_value >= column6_value:
            sheet.cell(row=42, column=4).fill = highlight
    else:
        print'NOT A NUMBER', (casted_value)


def get_ontario_category(CQAREF):
    array_values, _, _ = CQAUtilities.OntarioResults(CQAREF)
    Utilities.round_all_array_values(array_values)
    print('\nRESULTS FROM EXECUTION\n')
    check_array = Utilities.CQA_ON_DATA_CATEGORY
    check_array_second_part = Utilities.CQA_ON_SECOND_PART_CHECK

    CAT_A = False
    CAT_B = False
    CAT_FAIL = False

    # --FIRST SERIES OF CHECKS
    for i in array_values:
        # print(i[2])
        for j in check_array:
            if j[0] == i[1]:  # FIND THE INSTANCES IN THE ARRAYS WHEN THEY ARE THE SAME
                # print(i, j)
                if i[2] > j[3] and i[2] != 'BDL':  # check cat B
                    print(i, j, 'FAIL')
                    CAT_FAIL = True
                elif i[2] > j[2] and i[2] != 'BDL':  # check cat A
                    print(i, j, 'B')
                    CAT_B = True
                elif i[2] > j[1] and i[2] != 'BDL':  # check cat AA
                    print(i, j, 'A')
                    CAT_A = True
                else:
                    print(i, j, 'AA')
                    CAT_AA = True

        for k in check_array_second_part:
            if k[1] == i[1] and i[2] != 'NEGATIVE':
                # print(i, k)
                if i[2] > k[3] and i[1] != 'E. coli' and i[2] != 'BDL':
                    print(i, k, 'FAIL')
                    CAT_FAIL = True
                elif i[2] > k[2] and i[2] != 'BDL' and i[1] != 'E. coli':  # check cat A
                    print(i, k, 'B')
                    CAT_B = True
                else:  # check cat A
                    print(i, k, 'A')
                    CAT_A = True
                    CAT_AA = True

    print('CAT_A:     ' + str(CAT_A))
    print('CAT_B:     ' + str(CAT_B))
    print('CAT_FAIL:  ' + str(CAT_FAIL))

    if CAT_FAIL:
        return 'EXCEEDS GUIDELINES'
    elif CAT_B:
        return 'CATEGORY B'
    elif CAT_A:
        return 'CATEGORY A'


def get_non_ontario_category(CQAREF):
    array_values, _, _ = CQAUtilities.OntarioResults(CQAREF)
    Utilities.round_all_array_values(array_values)
    new_2d_array = array_values[0:11]
    check_array = Utilities.CQA_NON_ON_DATA_CATEGORY
    check_array_second_part = Utilities.CQA_NON_ON_SECOND_PART
    fecal_coli = Utilities.remove_greater_than(Utilities.get_fecal(CQAREF))
    print('\nRESULTS FROM EXECUTION\n')

    # for i in new_2d_array:
    #    print(i)
    print('checking values\n')
    CAT_A = False
    CAT_B = False
    CAT_FAIL = False

    for i in new_2d_array:
        # print(i[2])
        for j in check_array:
            if j[0] == i[1]:
                # print(i, j)
                if i[2] > j[2] and i[2] != 'BDL':  # check cat B
                    print(i, j, 'FAIL')
                    CAT_FAIL = True
                elif i[2] > j[1] and i[2] != 'BDL':  # check cat A
                    print(i, j, 'B')
                    CAT_B = True
                else:
                    print(i, j, 'A')
                    CAT_A = True

    for i in array_values:
        for k in check_array_second_part:
            if k[1] == i[1]:
                if i[2] > k[3] and i[2] != 'BDL':
                    print(i, k, 'FAIL')
                    CAT_FAIL = True
                elif i[2] > k[2] and i[2] != 'BDL':  # check cat A
                    print(i, k, 'B')
                    CAT_B = True
                else:  # check cat A
                    print(i, k, 'A')
                    CAT_A = True

    try:
        if float(fecal_coli) >= 1000:
            CAT_FAIL = True
            print('FECAL COLI ', fecal_coli, 'FAIL')
    except:
        print 'Either fecal coliform cannot be cast, or it is not bigger than 1000'

    print('CAT_A:     ' + str(CAT_A))
    print('CAT_B:     ' + str(CAT_B))
    print('CAT_FAIL:  ' + str(CAT_FAIL))

    if CAT_FAIL:
        return 'EXCEEDS GUIDELINES'
    elif CAT_B:
        return 'CATEGORY B'
    elif CAT_A:
        return 'CATEGORY A'
    # if CAT_AA:
    #    Utilities.write_to_csv('CAT AA')


#get_non_ontario_category('CQA2200124')
#get_ontario_category('CQA2200094')
