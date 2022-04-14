from openpyxl.styles import Border, Side
import SQL_CONNECTOR
import Colors
import shutil, os, sys
import openpyxl
from openpyxl.drawing.image import Image
import Utilities


def DQAcoverPageWrite(CQARef, workingFolder):
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    print Colors.bcolors.OKCYAN + "\nExecuting DQACover Page Write %s" % CQARef + Colors.bcolors.ENDC
    template_file = 'C:\CQA\FULL CQA - DQA\C&DQA\Templates\coverDQA.xlsx'
    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport", CQARef)
    dqaFile = saveLocation + r'\%sCoverDQA.xlsx' % (CQARef)
    cnx = SQL_CONNECTOR.test_connection()

    wb = openpyxl.load_workbook(template_file)
    if wb is None:
        print "Invalid Workbook"

    sheet = wb.get_sheet_by_name('CoverPage')

    # --
    sheet.cell(row=7, column=8).value = str(CQARef)
    sheet.cell(row=7, column=2).value = Utilities.get_company_name(CQARef)
    sheet.cell(row=8, column=2).value = Utilities.get_company_address(CQARef)
    sheet.cell(row=9, column=2).value = Utilities.get_FULL_ADDRESS(CQARef)
    sheet.cell(row=11, column=8).value = Utilities.get_sample_ID(CQARef)
    sheet.cell(row=13, column=2).value = Utilities.get_rptno(CQARef)
    sheet.cell(row=17, column=5).value = Utilities.full_location(CQARef)
    sheet.cell(row=18, column=2).value = Utilities.get_feecode(CQARef)
    sheet.cell(row=22, column=1).value = Utilities.get_sample_ID(CQARef)

    border = Border(bottom=thin)
    Utilities.FixFormatting(sheet, 'A36:D36', border)

    border = Border(bottom=thin)
    Utilities.FixFormatting(sheet, 'F36:I36', border)

    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')
    img = Image('hs.jpg')
    sheet.add_image(img, 'B35')
    img = Image('ian.bmp')
    sheet.add_image(img, 'H35')

    wb.save(dqaFile)
    cnx.close()
