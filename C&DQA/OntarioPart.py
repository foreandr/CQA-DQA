import os, sys, shutil
import mysql.connector
from mysql.connector import errorcode
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.styles import Font
import SQL_CONNECTOR
import Utilities
from Utilities import FixFormatting, config, formatDict, organicCarbon, calculateCEC, percentCalc, merge_two_dicts
import Colors


def sliceBeginningMaturity(string):

    if str(string[0]) == '<':
        return float(string[1:])
    else:
        return string

def slicePercentOffUnicode(string):
    if string == "BDL":
        return string
    if type(string) == int or type(string) == long:
        return string
    for i in string:
        if i == '%':
            string = string[:-1]
    return float(string)


def OntarioPrint(Workbook, finalResult, CQARef):
    print Colors.bcolors.OKCYAN + "\nExecuing Ontario Print" + Colors.bcolors.ENDC
    HOLDING_CATEGORY_STATE = {}
    # Runs the method to check what values have failed standards
    #highlightList = FailStandard(CQARef)
    #print 'HIGHLIGHT LIST [' + str(highlightList) + ']'
    # This is the sheet name inside of the template must be exact
    sheet = Workbook.get_sheet_by_name("CCME (Ontario)")
    # Sets the color of the highlight/fill to highlight the failed values
    highlight = PatternFill(start_color='F3F315', end_color='F3F315', fill_type='solid')

    # Unicode to add subscripts
    sheet['A33'] = u'CO\u2082 Respiration Rate'
    sheet['A34'] = u'CO\u2082 Respiration Rate'
    sheet['A35'] = u'O\u2082 Uptake Respiration Rate'
    sheet['A36'] = u'O\u2082 Uptake Respiration Rate'

    # Prints the value to the Excel sheet for each one by checking all the boxes that should have a number in them
    print Colors.bcolors.OKCYAN + "Ontario----Trace Elements----" + Colors.bcolors.ENDC
    # D10-20: 1-11
    for i in range(9, 20, 1):
        # finds what value is currently in the cell(A1 for example)
        value = sheet.cell(row=i, column=4).value

        try:
            # puts the value of the cell into string format
            cellName = str(value)
            # uses the value to search for the result
            valueToFill = finalResult[cellName]
        except KeyError:
            # If there isn't a result fill in a error
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        # Fill in the valueToFill variable
        sheet.cell(row=i, column=4).value = valueToFill

        print 'TESTING FOR CATEGORIES: ' + str(sheet.cell(row=i, column=4).value) + " & " + str(
            sheet.cell(row=i, column=6).value)

        print 'WHY CANT THIS BE CASTED' , sheet.cell(row=i, column=4).value
        if sheet.cell(row=i, column=4).value == 'BDL' or sheet.cell(row=i, column=4).value == 'Error':
            pass
        elif float(sheet.cell(row=i, column=4).value) > sheet.cell(row=i, column=6).value:
            sheet.cell(row=i, column=4).fill = highlight

        if sheet.cell(row=i, column=4).value == 'BDL' or sheet.cell(row=i, column=4).value == 'Error':
            sheet.cell(row=i, column=10).value = 'AA'
        elif float(sheet.cell(row=i, column=4).value) < sheet.cell(row=i, column=6).value:
            sheet.cell(row=i, column=10).value = 'AA'
        elif float(sheet.cell(row=i, column=4).value) < sheet.cell(row=i, column=7).value:
            sheet.cell(row=i, column=10).value = 'A'
        elif float(sheet.cell(row=i, column=4).value) < sheet.cell(row=i, column=8).value:
            sheet.cell(row=i, column=10).value = 'B'

        HOLDING_CATEGORY_STATE.update(
            {sheet.cell(row=i, column=2).value: sheet.cell(row=i, column=10).value}
        )

        '''--------------------HIGHLIGHTING RELEVANT COLUMNS-----------------------------------------'''

        # print "Current row: " + str(i) + " | " + highlightList[str(value)]  # what is the 0 and nothing representing

        # Check if the value fails standards if there is a none value for some reason skip it instead of crashing the program
        """
        if value == None:
            continue
        elif highlightList[
            str(value)] == '1':  # What does the 1 signify here? What is it checking? if I change it to 0? or != 0?
            sheet.cell(row=i, column=4).fill = highlight
        """
        '''-------------------------------------------------------------'''

    print Colors.bcolors.OKCYAN + "Ontario----foreign matter----" + Colors.bcolors.ENDC
    # D25 - 27: 39-40 & 12
    for i in range(24, 27, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=4).value = valueToFill
        """
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight
        """
        value_x = (sheet.cell(row=i, column=4).value)
        print '24-27 Value before slice : ' + str(value_x)
        value_x = slicePercentOffUnicode(value_x)
        print '24-27 Value after slice  : ' + str(value_x)

        if value_x == 'BDL':
            sheet.cell(row=i, column=10).value = 'AA/A'
        elif value_x < 1:
            sheet.cell(row=i, column=10).value = 'AA/A'
        elif value_x < 2:
            sheet.cell(row=i, column=10).value = 'B'
        elif value_x > 2:
            sheet.cell(row=i, column=10).value = 'Fail'

        HOLDING_CATEGORY_STATE.update(
            {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=10).value}
        )


    for i in range(28, 30, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        """
        sheet.cell(row=i, column=4).value = valueToFill
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight
        """

        value_x = (sheet.cell(row=i, column=4).value)
        print '28-29 Value before slice : ' + str(value_x)
        value_x = slicePercentOffUnicode(value_x)
        print '28-29 Value after slice  : ' + str(value_x)
        if value_x == 'BDL':
            sheet.cell(row=i, column=10).value = 'AA/A'
            pass
        elif value_x < 1:
            sheet.cell(row=i, column=10).value = 'AA/A'
            pass
        elif value_x < 2:
            sheet.cell(row=i, column=10).value = 'B'
            pass
        elif value_x > 2:
            sheet.cell(row=i, column=10).value = 'Fail'
            pass

        HOLDING_CATEGORY_STATE.update(
            {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=10).value}
        )

    print Colors.bcolors.OKCYAN + "Ontario----Maturity/Stability----" + Colors.bcolors.ENDC
    # D34: 15
    value = sheet.cell(row=33, column=4).value
    try:
        cellName = str(value)
        valueToFill = finalResult[cellName]
    except KeyError:
        valueToFill = "Error"
        sheet.cell(row=i, column=4).fill = highlight

    sheet.cell(row=33, column=4).value = valueToFill
    """
    if value == None:
        print ''
    elif highlightList[str(value)] == '1':
        sheet.cell(row=33, column=4).fill = highlight
    """

    test_value = Utilities.getCO2Resp(CQARef)
    print '33 CO2 VALUE: ' + str(test_value)
    #print (type(test_value))
    if float(test_value) > 4:
        sheet.cell(row=33, column=10).value = 'Fail'
        pass
    else:
        sheet.cell(row=33, column=10).value = 'AA'
        pass

    HOLDING_CATEGORY_STATE.update(
        {sheet.cell(row=33, column=1).value: sheet.cell(row=33, column=10).value}
    )

    print Colors.bcolors.OKCYAN + "Ontario----Pathogens----" + Colors.bcolors.ENDC
    # D43-44: 16-17
    for i in range(40, 42, 1):
        value = sheet.cell(row=i, column=4).value
        # print value
        try:
            cellName = str(value)
            # print cellName
            valueToFill = finalResult[cellName]
            # print valueToFill
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=4).value = valueToFill
        """
        if value == None:
            continue
        elif highlightList[str(value)] == '1':
            sheet.cell(row=i, column=4).fill = highlight
        """


        # print '40-42: ' + valueToFill

        # CHECKING 40
        if i == 40:
            tester = sliceBeginningMaturity(valueToFill)
            if tester < 1000:
                sheet.cell(row=i, column=10).value = 'AA'
                pass
            else:
                sheet.cell(row=i, column=10).value = 'Fail'
                pass


        # CHECKING 41
        if i == 41:
            tester = sliceBeginningMaturity(valueToFill)
            if tester == "NEGATIVE":
                sheet.cell(row=i, column=10).value = 'Cat-AA'
                pass
            else:
                sheet.cell(row=i, column=10).value = 'Fail'
                pass

        HOLDING_CATEGORY_STATE.update(
            {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=10).value}
        )

    print Colors.bcolors.OKCYAN + "------------------" + Colors.bcolors.ENDC
    for i in HOLDING_CATEGORY_STATE:
        print i, HOLDING_CATEGORY_STATE[i]
    #print (len(HOLDING_CATEGORY_STATE))
    # remove extra values from excel
    for i in range(1, 50):
        sheet.cell(row=i, column=10).value = None


    print  Colors.bcolors.OKGREEN +'Category of %s refno is: '%CQARef + Utilities.return_worst(HOLDING_CATEGORY_STATE) + Colors.bcolors.ENDC

    print Colors.bcolors.OKCYAN + "------------------\n" + Colors.bcolors.ENDC

    print Colors.bcolors.OKCYAN + "Ontario----CFIA----" + Colors.bcolors.ENDC
    # F54-55: 18-19
    for i in range(46, 48, 1):
        value = sheet.cell(row=i, column=6).value  # THIS IS GRABBING THE VALUE FROM THE TEMPLATE
        try:
            print 'value -----------' + str(value)
            cellName = str(value)  # cell name is 18 | then 19, total organic matter index
            print 'cell name -----------' + cellName
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        '''TOTAL ORGANIC MATTER LOCATION'''
        sheet.cell(row=i, column=6).value = valueToFill
        # sheet.cell(row=i, column=6).value = "Andre"  #

    print Colors.bcolors.OKCYAN + "Ontario----Finished Compost Quality----" + Colors.bcolors.ENDC
    # f60-64: 20-24
    for i in range(53, 58, 1):
        value = sheet.cell(row=i, column=6).value
        try:
            print '53-58 value -----------' + str(value)
            cellName = str(value)
            print '53-58 cell name -----------' + cellName
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=6).value = valueToFill

    # f66-68: 25-27
    for i in range(59, 62, 1):
        value = sheet.cell(row=i, column=6).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]

        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight

        sheet.cell(row=i, column=6).value = valueToFill

    print Colors.bcolors.OKCYAN + "Ontario----Compost Agricultural Product Value----" + Colors.bcolors.ENDC
    # D96-99: 28-31
    for i in range(100, 104, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight
        print i
        print valueToFill
        print value
        sheet.cell(row=i, column=4).value = valueToFill

    # D101-107: 32-38
    for i in range(105, 112, 1):
        value = sheet.cell(row=i, column=4).value
        try:
            cellName = str(value)
            valueToFill = finalResult[cellName]
        except KeyError:
            valueToFill = "Error"
            sheet.cell(row=i, column=4).fill = highlight
        print i
        print valueToFill
        print value
        sheet.cell(row=i, column=4).value = valueToFill

    # Below is all the code used to fix the borders by giving it a range of cell values(eg. A10:I10) and using Border to change the thickness size
    # Creates the border styles for the different side types
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    # Trace metals
    # This line sets the border values
    border = Border(top=thick, left=thick, right=thick, bottom=thick)
    # This line calls the FixFormatting function

    FixFormatting(sheet, 'B6:D6', border)

    # This changes the border for the one needed and it just kinda keeps rotating

    ####bottom border of the table "A. Maximum Concentrations for Trace Metals in Compost Ontario"
    border = Border(bottom=thick)
    FixFormatting(sheet, 'B19:D19', border)

    # for the border box "Maximum Concentration within Product"
    border = Border(right=thick)
    FixFormatting(sheet, 'E7:H7', border)
    FixFormatting(sheet, 'E8:H8', border)

    border = Border(bottom=thick)
    FixFormatting(sheet, 'E8:H8', border)

    # Category AA, A, B Box border
    border = Border(bottom=thick, top=thick)
    FixFormatting(sheet, 'F6:H6', border)

    border = Border(bottom=thick)
    FixFormatting(sheet, 'F19:G19', border)  # can be combined with the code ("B19:D19")

    # border = Border(top=thick, left=thick, right=thick, bottom=thick)
    # FixFormatting(sheet, 'B6:C6',border)

    # This is for the walls of the borders where i can just run a for loop instead of doing it manually

    #####Test Results and Category AA Column######
    border = Border(right=thin, bottom=thin)
    for i in range(9, 19):
        cellNumber = 'B%s:C%s' % (i, i)  # Test Results and Category AA Column
        FixFormatting(sheet, cellNumber, border)

    #####Category A and B Column right borders######
    border = Border(bottom=thin)
    for i in range(9, 19):
        cellNumber = 'F%s:G%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # Foreign Matter
    ####Bottom Border####
    border = Border(bottom=thick)
    FixFormatting(sheet, 'A29:I29', border)

    # right bottom border of Table B
    border = Border(bottom=thick, right=thick)
    FixFormatting(sheet, 'H29:I29', border)

    #####Column Heading row borders (Test Results, Category AA, A, B)#####
    border = Border(top=thick, left=thin, right=thick, bottom=thick)
    FixFormatting(sheet, 'A22:I22', border)

    # Foreign matter header boxes#
    border = Border(top=thick, left=thin, right=thin, bottom=thin)
    FixFormatting(sheet, 'E23:G26', border)
    border = Border(top=thick, left=thin, right=thick, bottom=thin)
    FixFormatting(sheet, 'H23:I26', border)
    border = Border(top=thin, left=thin, right=thin, bottom=thick)
    FixFormatting(sheet, 'E27:G29', border)
    # boxes in column D sharps
    border = Border(top=thin, right=thin, bottom=thin)
    FixFormatting(sheet, 'A27:D27', border)
    # boxes in column D fm
    border = Border(top=thick, right=thin, bottom=thin)
    FixFormatting(sheet, 'A23:D23', border)

    '''MY BORDER FIXES'''

    #####Foregin Matter Explanation Box#####
    border = Border(right=thin, bottom=thin)
    for i in range(23, 29):
        cellNumber = 'B%s:C%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    #####Right border for cells under category B column#####
    border = Border(right=thick)
    for i in range(22, 29):
        cellNumber = 'H%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # Maturity and stability
    # First row of Table C
    border = Border(top=thick, bottom=thick)
    FixFormatting(sheet, 'A32:I32', border)

    #####Right border of Table C#####
    border = Border(right=thick, left=thick, bottom=thick)
    FixFormatting(sheet, 'A32:I36', border)

    #####Border for the method table cell#####
    border = Border(right=thin, bottom=thin)
    for i in range(33, 36, 2):
        cellNumber = 'A%s:C%s' % (i, i + 1)
        FixFormatting(sheet, cellNumber, border)

    #####Border for Required Limits entry box#####
    border = Border(left=thin, bottom=thin, right=thin)
    for i in range(33, 36, 2):
        cellNumber = 'E%s:I%s' % (i, i + 1)
        FixFormatting(sheet, cellNumber, border)

    # Pathogens
    #####Top border of Pathogen box#####
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A39:I39', border)

    #####Second Row of Pathogen#####
    border = Border(right=thick, bottom=thin)
    FixFormatting(sheet, 'A40:I40', border)

    #####Third Row of Pathogen#####

    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'A41:I41', border)

    # CFIA
    #####Top Border of CIFA Table#####
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'C45:G45', border)

    #####Right and Bottom for Organic Matter#####
    border = Border(right=thick, left=thin, bottom=thin)
    FixFormatting(sheet, 'F46:G46', border)

    #####Right and Bottom for Moisture#####
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'F47:G47', border)

    # Compost Quality
    #####Compost table column heading row#####
    border = Border(right=thick, bottom=thick, top=thick)
    FixFormatting(sheet, 'C52:G52', border)

    #####Compost table bottom row#####
    border = Border(right=thick, bottom=thick)
    FixFormatting(sheet, 'C61:G61', border)

    #####result cells for Compost Quality#####
    border = Border(right=thick, bottom=thin)
    for i in range(53, 61):
        cellNumber = 'C%s:G%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    # reference Compost Quality

    '''FIXED SOME FORMATTING ERRORS HERE'''
    border = Border(bottom=thick, top=thick)
    FixFormatting(sheet, 'A72:I72', border)

    '''FIXED SOME FORMATTING ERRORS HERE'''
    border = Border(top=thick, bottom=thin)
    FixFormatting(sheet, 'G83:H83', border)

    '''FIXED SOME FORMATTING ERRORS HERE'''
    border = Border(bottom=thin)
    for i in range(72, 85):  # Andre: Changed from 75 to 82 to accomodate, not sure why it was 82
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    '''FORMATTING ERRORS LINE 112 RIGHTSIDE'''
    border = Border(right=thick)
    FixFormatting(sheet, 'I112:I112', border)

    '''FORMATTING ERRORS LINE 112'''
    border = Border(bottom=thick, top=thin)
    FixFormatting(sheet, 'A112:I112', border)

    '''FORMATTING ERRORS LINE 113'''
    border = Border(bottom=thick, top=thin)
    FixFormatting(sheet, 'A113:I113', border)

    '''FIXED SOME FORMATTING ERRORS HERE'''
    border = Border(right=thick)
    FixFormatting(sheet, 'I113:I113', border)

    # Compost product value Top line
    border = Border(bottom=thick, top=thick, right=thick)
    FixFormatting(sheet, 'A98:I98', border)

    border = Border(right=thick)
    FixFormatting(sheet, 'A98:I111', border)

    border = Border(top=thin, right=thick, bottom=thin)
    for i in range(100, 111):  ##98, 110
        cellNumber = 'A%s:I%s' % (i, i)
        FixFormatting(sheet, cellNumber, border)

    border = Border(bottom=thin)
    FixFormatting(sheet, 'A114:I114', border)

    # imports all the photos to the sheet
    from openpyxl.drawing.image import Image
    # The directory of where the photos are kept
    os.chdir(r'C:\CQA\NewFormatCQA\Photos')
    # AL
    # Creating an image variable
    img = Image('al.jpg')
    # Adding it to the sheet at A1
    sheet.add_image(img, 'A1')
    img2 = Image('al.jpg')
    sheet.add_image(img2, 'A49')
    img3 = Image('al.jpg')
    sheet.add_image(img3, 'A94')

    # Cp
    img = Image('cp.png')
    sheet.add_image(img, 'H1')
    img = Image('cp.png')
    sheet.add_image(img, 'H49')
    img = Image('cp.png')
    sheet.add_image(img, 'H94')

    '''OG AGINDEX PICTURE HAD ISSUES'''

    font_c = Font(color='000000', size=10)
    # sheet.cell(row=115, column=1).value = "salt injury probably"
    # sheet.cell(row=115, column=3).value = "apply on salts with excellent drainage chracteristics good water quality and low salts"
    # sheet.cell(row=115, column=6).value = "apply on salts with poor drainage poor water quality, or high salts"
    # sheet.cell(row=115, column=9).value = "for all salts"
    for i in range(1, 10):  # Changing color from red to black
        sheet.cell(row=115, column=i).font = font_c

    # sheet.cell(row=116, column=5).value = "fig 05.02-F1 AgIndex interpretation and use for commmon edaphic conditions"
    # sheet.cell(row=116, column=5).font = font_c
    '''DOING CALCULATION FOR AGINDEX-------------------------------------------------------------------------------------------'''
    #
    value = 'TEMP'
    N = sheet.cell(row=105, column=4).value  # TOTAL NITROGEN
    P2O5 = sheet.cell(row=107, column=4).value  # PHOSPHATE
    K20 = sheet.cell(row=108, column=4).value  # POTASSIUM

    cnx = SQL_CONNECTOR.test_connection()
    cursor = cnx.cursor()
    query = """
        select rep.rptno, rep.refno, env.feecode, env.result
        from alms.report rep
        inner join alms.env_data env
        on rep.rptno = env.rptno
        where env.feecode = 'mnac380'
        and rep.refno = '%s' 
        """ % CQARef
    cursor.execute(query)
    NA = ''
    for i in cursor:
        NA = i[3]

    cursor = cnx.cursor()
    CL = ''
    query = """
        select soil.cl
        from alms.report rep
        inner join alms.soil soil
        on rep.rptno = soil.rptno
        where rep.refno = '%s'
    """ % CQARef
    cursor.execute(query)
    for i in cursor:
        # print i
        CL = i[0]

    initialUnicodeList = [N, P2O5, K20, NA, CL]
    updatedFloatList = []
    for i in initialUnicodeList:
        i = str(i)
        i = slicePercentOffUnicode(i)
        updatedFloatList.append(i)

    for i in updatedFloatList:
        # print i
        # print type(i)
        pass

    drymatter = sheet.cell(row=100, column=4).value  # TOTAL NITROGEN

    # (dry matter /100 )* sodium
    Nitrogen = updatedFloatList[0]
    Phosphorus = updatedFloatList[1]
    Potassium = updatedFloatList[2]
    Sodium = updatedFloatList[3]
    Chloride = updatedFloatList[4]
    DryMatter = slicePercentOffUnicode(str(drymatter))

    a_index = (Nitrogen + Phosphorus + Potassium) / ((Sodium * (DryMatter / 100)) + (Chloride / 10000))
    sheet.cell(row=113, column=4).value = round(a_index)  # cast to integer

    # Putting in the comment
    print decisionAgIndex(a_index)
    font_c = Font(color='000000', size=10)
    sheet.cell(row=113, column=6).value = decisionAgIndex(a_index)
    sheet.cell(row=113, column=6).font = font_c

    # AGINDEX ADDITION
    img = Image("C:/CQA/FULL CQA - DQA/C&DQA/Photos/agindex.png")
    sheet.add_image(img, 'b115')

    '''FINISHED CALCULATION FOR AGINDEX-------------------------------------------------------------------------------------------'''

    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport", CQARef)

    Workbook.save(saveLocation + "\%sReport.xlsx" % (CQARef))


def decisionAgIndex(num):
    if num == 2:
        return 'Salt Injury Probable'
    elif num >= 2 and num <= 5:
        return 'Apply on soils with excellent drainage characteristics, good water quality and low salts'
    elif num >= 5 and num <= 10:
        return 'Apply on salts with poort drainage, poor water quaity or high salts'
    elif num > 10:
        return 'for all salts'


def findOntarioCatagory(CQARef):
    """makes a list and adds the values to the list and checks what category they're in, returns the category letter"""
    # gets the results from the Ontario Results Function
    ENVResult = OntarioResults(CQARef)
    # Makes a dictionary for the results needed(1-17 are needed)
    results = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '40': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # Puts the 17 results needed into a separate list called valueCatList
    valueCatList = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '40': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # The max values for each category(Arsenic catAA=13 catA=13 catB=75) -1 means there is no value in that category and will be skipped
    # If there is a CatAA and a CatB but no Cat A then put cat A the same as catAA or else it will be put down as exceeds if you use -1
    catAA = {
        '1': 13,
        '2': 3,
        '3': 210,
        '4': 34,
        '5': 100,
        '6': 150,
        '7': 0.8,
        '8': 5,
        '9': 62,
        '10': 2,
        '11': 500,
        '39': 1,
        '40': 1,
        '12': 1,
        '13': 0,
        '14': 0,
        '15': 4,
        '16': 1000,
        '17': 3
    }
    catA = {
        '1': 13,
        '2': 3,
        '3': 210,
        '4': 34,
        '5': 400,
        '6': 150,
        '7': 0.8,
        '8': 5,
        '9': 62,
        '10': 2,
        '11': 700,
        '39': 1,
        '40': 1,
        '12': 1,
        '13': 0,
        '14': 0,
        '15': -1,
        '16': -1,
        '17': -1
    }

    catB = {
        '1': 75,
        '2': 20,
        '3': 1060,
        '4': 150,
        '5': 760,
        '6': 500,
        '7': 5,
        '8': 20,
        '9': 180,
        '10': 14,
        '11': 1850,
        '39': 2,
        '40': 2,
        '12': 2,
        '13': 3,
        '14': 3,
        '15': -1,
        '16': -1,
        '17': -1
    }
    # The category defaults to AA
    finalResult = 'AA'

    # Stores the values into valueCatList
    for i in results.keys():
        if i not in valueCatList:
            continue
        valueCatList[i] = ENVResult[i]

    for key in results.keys():

        # Temp is used to store and modify the current value its on to remove percent signs
        temp = None
        try:
            temp = valueCatList[key].replace('%', '')
        # If the value is an integer then dont bother doing it
        except AttributeError:
            temp = valueCatList[key]

        # Make sure that all the values that can be numbers are numbers
        try:
            value = float(temp)
        # If it can't be a number then don't change anything
        except ValueError:
            value = valueCatList[key]

        '''ADDITION BY ANDRE'''
        print 'CURRENT KEY:VALUE == ' + str(key) + ': ' + str(value)

        # If the value is BDL then it has to be AA
        if value == 'BDL*':
            results[key] = 'AA'

        # If the value is a string then find out what one it is and find the category by its type
        elif type(value) == str:
            # If the value of salmonella is negative then its AA
            if value == 'Negative' and key == '17' or value == 'NEGATIVE' and key == '17':
                results[key] = 'AA'
            # If the value of salmonella is positive than it exceeds
            elif value == 'Positive' and key == '17' or value == 'POSITIVE' and key == '17':
                results[key] = 'Exceeds'
            # If its E. Coli and its <3 then its AA
            elif value == '<3' and key == '16':
                results[key] = 'AA'
            # if its E. Coli and its >1000 then it exceeds
            elif value == '>1000' and key == '16':
                results[key] = 'Exceeds'
            # if its Ecoli and its <3 BUT also salmonella is positive than it exceeds
            elif value == '<3' and key == '16' and value == 'Positive' and key == '17':
                results[key] = 'Exceeds'

        # If the value is less then category AA then its AA
        elif value <= catAA[key]:
            results[key] = 'AA'

        elif value > catAA[key]:
            # If it's not AA and Category A doesn't exist then it Exceeds
            if catA == -1:
                results[key] = 'Exceeds'

            else:
                # If Category A exists and the value is lower than it than it's A
                if value <= catA[key]:
                    results[key] = 'A'

                else:
                    # If its above category A and Category B doesn't exist then it Exceeds
                    if catB[key] == -1:
                        results[key] = 'Exceeds'

                    else:
                        # If the value is below category B then its B
                        if value <= catB[key]:
                            results[key] = 'B'
                        # If it doesn't fit in any of the above things then it Exceeds
                        else:
                            results[key] = 'Exceeds'

    # goes through all the dictionary results and finds the highest category letter and returns that one
    for key in results.keys():
        parameter = results[key]

        if finalResult == parameter:
            finalResult = parameter
        elif finalResult == 'AA' and (parameter == 'A' or parameter == 'B' or parameter == 'Exceeds'):
            finalResult = parameter
        elif finalResult == 'A' and (parameter == 'B' or parameter == 'Exceeds'):
            finalResult = parameter
        elif finalResult == 'B' and parameter == 'Exceeds':
            finalResult = parameter

    return finalResult


def FailStandard(CQARef):
    results = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '40': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # gets the results from the Ontario Results Function
    ENVResult = OntarioResults(CQARef)
    # Puts the 17 results needed into a separate list called valueCatList
    valueCatList = {
        '1': '',
        '2': '',
        '3': '',
        '4': '',
        '5': '',
        '6': '',
        '7': '',
        '8': '',
        '9': '',
        '10': '',
        '11': '',
        '39': '',
        '40': '',
        '12': '',
        '13': '',
        '14': '',
        '15': '',
        '16': '',
        '17': ''
    }
    # category max
    catMax = {
        '1': 75,
        '2': 20,
        '3': 1060,
        '4': 150,
        '5': 760,
        '6': 500,
        '7': 5,
        '8': 20,
        '9': 180,
        '10': 14,
        '11': 1850,
        '39': 2,
        '40': 2,
        '12': 2,
        '13': 3,
        '14': 3,
        '15': 4,
        '16': 1000,
        '17': 3
    }
    # Stores the values into valueCatList
    for key in valueCatList.keys():
        valueCatList[key] = ENVResult[key]

    for key in valueCatList.keys():
        # Temp is used to store and modify the current value its on to remove percent signs
        temp = None
        try:
            temp = valueCatList[key].replace('%', '')
        # If the value is an integer then dont bother doing it
        except AttributeError:
            temp = valueCatList[key]

        # Make sure that all the values that can be numbers are numbers
        try:
            value = float(temp)
        # If it can't be a number then don't change anything
        except ValueError:
            value = valueCatList[key]

        # If the value is BDL then it has to be 0
        if value == 'BDL':
            results[key] = '0'

        # If the value is a string then find out what one it is and find the category by its type
        elif type(value) == str:
            # If the value of salmonella is negative then its AA
            if value == 'Negative' and key == '17' or value == 'NEGATIVE' and key == '17':
                results[key] = '0'
            # If the value of salmonella is positive than it exceeds
            elif value == 'Positive' and key == '17' or value == 'POSITIVE' and key == '17':
                results[key] = '1'
            # If its E. Coli and its <3 then its AA
            elif value == '<3' and key == '16':
                results[key] = '0'
            # if its E. Coli and its >1000 then it exceeds
            elif value == '>1000' and key == '16':
                results[key] = '1'
        elif value >= catMax[key]:
            results[key] = '1'
        else:
            results[key] = '0'

    return results


def OntarioResults(CQARef):
    print Colors.bcolors.OKCYAN + "\nExecuting Ontario Results" + Colors.bcolors.ENDC
    # ------------------------------------------------------Env Report------------------------------------------------------#
    # Dictionary for the environmental report values
    ENVDict = {
        "1": "Arsenic",
        "2": "Cadmium",
        "3": 'Chromium',
        "4": 'Cobalt',
        "5": 'Copper',
        '6': 'Lead',
        "7": 'Mercury',
        '8': 'Molybdenum',
        '9': 'Nickel',
        '10': 'Selenium',
        '11': 'Zinc',
        '12': 'Total FM > 25 mm',
        '13': 'Total sharps > 2.8 mm*',
        '14': 'Total sharps > 12.5 mm',
        '15': 'Respiration-mgCO2-C/g OM/day',
        '16': 'E. coli',
        '17': 'Salmonella spp.',
        '18': 'Total Organic Matter',
        '19': 'Moisture',
        '20': 'Total Organic Matter @ 550 deg C',
        '22': 'C:N Ratio',
        '28': 'Total Solids (as received)',
        '30': 'Bulk Density (As Recieved)',
        '33': 'Ammonia (NH3/NH4-N)',
        '34': 'Total Phosphorus (As P205)',
        '35': 'Total Potassium (as K20)',
        '36': 'Calcium',
        '37': 'Magnesium',
        '38': 'Sulphur',
        '39': 'Total FM > 2.8 mm*',
        '40': 'Total plastics > 2.8 mm*'
    }

    # CoNnEcT To dB
    cnx = SQL_CONNECTOR.test_connection()

    # Querry out the report numbers
    cursor = cnx.cursor()
    query = "SELECT * FROM alms.report WHERE refno='%s'" % (CQARef)
    cursor.execute(query)
    reportNumbers = []
    for item in cursor:
        print Colors.bcolors.HEADER + "[Ontario Report REF Query:]" + Colors.bcolors.ENDC + str(item)
        tempList = [item[0], item[1]]
        reportNumbers.append(tempList)

    # Store the report numbers
    soilReport = None
    envReport = None
    for item in reportNumbers:
        if str(item[1]) == "SOIL":
            soilReport = str(item[0])
        elif str(item[1]) == "ENVI":
            envReport = str(item[0])

    print soilReport, envReport

    ENVResult = {}

    for key in ENVDict.keys():
        parameter = ENVDict[key]
        # query every parameter based on name
        if parameter != "Total FM > 25 mm":
            print key, parameter
            envQuery = r"""select
            ed.result_str
            from env_data as ed
            inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
            inner join report as r on ed.rptno=r.rptno
            inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
            left join units as u on ed.unit=u.un_units
            where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='%s'
            order by f.prt_sort,f.anly_sort""" % (envReport, parameter)

        # 12
        # query for specific unit of Total FM >25 mm parameter
        elif parameter == "Total FM > 25 mm":
            envQuery = r"""select
                ed.result_str
                from env_data as ed
                inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
                inner join report as r on ed.rptno=r.rptno
                inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
                left join units as u on ed.unit=u.un_units
                where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.unit = "pieces/500ml" and ed.rptno='%s' and f.rpt_dscrpt='%s'
                order by f.prt_sort,f.anly_sort""" % (envReport, parameter)

        cursor = cnx.cursor()
        cursor.execute(envQuery)
        # Store the querried information in the ENVResult list
        for item in cursor:
            ENVResult[key] = str(item[0])
        print Colors.bcolors.BOLD + Colors.bcolors.UNDERLINE + Colors.bcolors.OKBLUE + \
              "Key: " + str(key) + " | " + \
              "Parameter: " + str(parameter) \
              + "" + Colors.bcolors.ENDC

    print Colors.bcolors.OKCYAN + 'Ontario----salmonella------' + Colors.bcolors.ENDC
    print ENVResult['17']
    ##    print '=-----moisture------'
    ##    print ENVResult['19']

    # Querry out moisture to calculate Results as recieved
    envQuery = r"""select
            ed.result
            from env_data as ed
            inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
            inner join report as r on ed.rptno=r.rptno
            inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
            left join units as u on ed.unit=u.un_units
            where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='Moisture'
            order by f.prt_sort,f.anly_sort""" % (envReport)

    cursor = cnx.cursor()
    cursor.execute(envQuery)
    # Store the querried information
    for item in cursor:
        moisture = item[0]
        ENVResult['19'] = moisture

    for i in range(34, 39):
        temp = str(i)
        dm = 100 - float(moisture)
        dm = dm / 100
        ENVResult[temp] = str(float(ENVResult[temp]) * dm)
        # ------------------------------------------------------Sieves------------------------------------------------------#
    # 22
    # List of all possible Sieve sizes
    sieveList = [r'Sieve 2 Inch (% Passing)', r'Sieve 1 Inch (% Passing)', r'Sieve 1/2 Inch (% Passing)',
                 r'Sieve 3/8 Inch (% Passing)',
                 r'Sieve 1/4 Inch (% Passing)']

    sieveDict = {}
    sieveResult = {}
    targetPercent = 80.0
    smallestParameter = None
    # Querry out the sieve sizes
    for item in sieveList:
        parameter = item
        envQuery = r"""select
        ed.result_str
        from env_data as ed
        inner join env_samp as es on ed.rptno=es.rptno and ed.labno=es.labno
        inner join report as r on ed.rptno=r.rptno
        inner join feecode as f on ed.feecode=f.feecode and left(f.rpt_dscrpt,4)<>'TCLP' and f.rpt_flag<>'N'
        left join units as u on ed.unit=u.un_units
        where (ed.result>=0 or ed.result=-3 or ed.result_str='BDL*') and ed.rptno='%s' and f.rpt_dscrpt='%s'
        order by f.prt_sort,f.anly_sort""" % (envReport, parameter)
        cursor = cnx.cursor()
        cursor.execute(envQuery)
        # Goes through all sieve sizes
        for item in cursor:
            a = str(item[0])
            item = float(item[0])
            # Check to see if the current sieve size is below the target size and if it is then sets the value too 999
            if item - targetPercent <= 0:
                sieveDict[parameter] = 999
                sieveResult[parameter] = a
                # If its not below the target size then sets the value to the difference of the two numbers
            else:
                sieveDict[parameter] = item - targetPercent
                sieveResult[parameter] = a

    # get smallest size
    for key, value in sorted(sieveDict.iteritems(), key=lambda (k, v): (v, k)):
        smallestParameter = key
        break

    tempSieve = smallestParameter.replace('Sieve', '').replace('(% Passing)', '')[1:-1]

    # sets the sieve size
    ENVResult['22'] = tempSieve

    # ------------------------------------------------------Soil Report------------------------------------------------------#

    soilResult = {}

    # Querrys out the different soil report values
    # 20 and 29: PH
    soilResult['20'] = Utilities.getPH(CQARef)
    soilResult['29'] = Utilities.getPH(CQARef)

    # 23: salt
    query = """select
        IF(s.salt>0,s.salt,null) as salt
        from soil as s
        inner join report as r on r.rptno=s.rptno
        where s.rptno='%s'
        order by s.labno""" % (soilReport)

    cursor = cnx.cursor()
    cursor.execute(query)
    for item in cursor:
        soilResult['23'] = str(round(item[0], 1))

    soilResult['32'] = Utilities.getNitrogen(CQARef)

    # -----------Total Organic matter-------------#
    print Colors.bcolors.OKCYAN + "----------------------------------Ontario total organic-------------------------------" + Colors.bcolors.ENDC

    soilResult['32'] = Utilities.getNitrogen(CQARef)  # Nitrogen
    soilResult['18'] = Utilities.getTotalOrganicMatter(CQARef)  # Total organic MAtter
    available_matter_for_calc = Utilities.getAvailableOrganicMatter(CQARef)
    totalOrganicCarbon2 = organicCarbon(available_matter_for_calc)
    Nitrogen = float(soilResult['32'])
    print 'Total Organic            :' + str(soilResult['18'])
    print 'AVailable Organic        : ' + str(available_matter_for_calc)
    print 'OG CARB * 0.6            : ' + str(totalOrganicCarbon2)
    print 'Nitrogen                 :' + str(Nitrogen)

    # Divide organic carbon by nitrogen
    CNRatioValue = round((organicCarbon(available_matter_for_calc) / 0.9) / Nitrogen)
    print 'CNRatioValue             :' + str(CNRatioValue)

    cNRatio = str("%d:1" % (CNRatioValue))
    print 'Calculated CN Ratio = ' + cNRatio
    soilResult['21'] = cNRatio
    soilResult['31'] = cNRatio

    # ----------------------------Interpretation--------------------------------------------#

    # CEC Calculations (k_m3, mg_m3, ca_m3, na, buffer)
    # 24-27
    CECDict = {
        'k_m3': [],
        'mg_m3': [],
        'ca_m3': [],
        'na': [],
        'buffer': []
    }
    # Querry the cecDict
    for key in CECDict.keys():
        parameter = key
        query = """select
            IF(s.%s>0,s.%s,null) as %s
            from soil as s
            inner join report as r on r.rptno=s.rptno
            where s.rptno='%s'
            order by s.labno""" % (parameter, parameter, parameter, soilReport)

        cursor = cnx.cursor()
        cursor.execute(query)
        # If they don't exist then stop else add them to the dict
        for value in cursor:
            if value is None:
                sys.exit()
            else:
                CECDict[key].append(float(value[0]))

    # perk - k_m3, 390, E28
    # perMG - _mg_m3, 121.6, E29
    # perCa - ca_m3, 200.0, E30
    # perNa - na, 230.0., E26

    # Call the cakculateCEC function
    cec = calculateCEC(CECDict)

    # calculation parameters
    interpDict = {
        '24': ['na', 230.0],
        '25': ['k_m3', 390.0],
        '26': ['mg_m3', 121.6],
        '27': ['ca_m3', 200.0]
    }

    # add results to soilResults
    for key in interpDict.keys():
        # call def percentCalc(cec, paramterName, y, CECDict):
        result = percentCalc(cec, CECDict[interpDict[key][0]][0], interpDict[key][1], CECDict)
        soilResult[key] = result

        # ----------------------------Merging and Formatting--------------------------------------------#

    # Runs function that merges the two dict's
    tempResult = merge_two_dicts(ENVResult, soilResult)

    finalResult = {}
    # Goes through all the results
    for key in tempResult.keys():
        # Stores results into lists
        value = tempResult[key]
        digits = formatDict[key]

        # Sees what format the number needs and modifies the value to fit the format
        try:
            float(value)
            if digits == 2:
                finalValue = "%.2f" % (float(value))
                # print key, value, finalValue
                finalResult[key] = finalValue
            elif digits == 1:
                finalValue = "%.1f" % (float(value))
                # print key, value, finalValue
                finalResult[key] = finalValue
            elif digits == 0:
                finalValue = int(value)
                finalResult[key] = finalValue
            elif digits == '2%':
                finalValue = "%.2f" % (float(value)) + '%'
                finalResult[key] = finalValue

        # If there's a value error then don't change it
        except ValueError:
            finalResult[key] = value

    cursor.close()
    cnx.close()
    print 'Ontario Final Result ' + str(finalResult)
    return finalResult
