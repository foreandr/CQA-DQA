import os
from openpyxl.styles import Side, Border
import Colors
import Utilities


def CFIAPrintDQA(Workbook, CQARef):
    print(Workbook, CQARef)
    sheet = Workbook.get_sheet_by_name("CCME CFIA", )
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    import gettingFeeCodes

    feecodeList = gettingFeeCodes.gettingfeeCodes('%s' % CQARef)

    # else:
    #    print(i[1])
    # print('printing list for entering')
    # for i in list_for_entering:
    #    print i

    for i in feecodeList:
        print Colors.bcolors.OKGREEN + str(i) + Colors.bcolors.ENDC
    # ---------- CREATING CALC DICT
    dict_names_values = {}
    for i in feecodeList:
        dict_names_values[i[1]] = (i[3])
    typcasted_dict = {}
    for key, value in dict_names_values.iteritems():
        try:
            typcasted_dict[key] = float(value)
        except:
            # print(value, 'not a float, entering into dict as string')
            typcasted_dict[key] = value

    for key, value in typcasted_dict.items():
        print key, value

    dry_matter = typcasted_dict['Dry Matter']
    # print dry_matter
    # ---------
    # A
    sheet.cell(row=11, column=4).value = feecodeList[15][3]
    sheet.cell(row=12, column=4).value = feecodeList[18][3]
    sheet.cell(row=13, column=4).value = feecodeList[19][3]
    sheet.cell(row=14, column=4).value = feecodeList[20][3]
    sheet.cell(row=15, column=4).value = feecodeList[21][3]
    sheet.cell(row=16, column=4).value = feecodeList[30][3]
    sheet.cell(row=17, column=4).value = feecodeList[23][3]
    sheet.cell(row=18, column=4).value = feecodeList[27][3]
    sheet.cell(row=19, column=4).value = feecodeList[29][3]
    sheet.cell(row=20, column=4).value = feecodeList[32][3]
    sheet.cell(row=21, column=4).value = feecodeList[34][3]

    # B. Foreign Matter
    sheet.cell(row=28, column=4).value = feecodeList[36][3]
    sheet.cell(row=29, column=4).value = feecodeList[37][3]
    sheet.cell(row=30, column=4).value = feecodeList[35][3]

    # D Pathogens
    sheet.cell(row=34, column=4).value = feecodeList[2][3]
    sheet.cell(row=35, column=4).value = feecodeList[3][3]

    # Minimum Agricultural Values
    sheet.cell(row=52, column=4).value = feecodeList[11][3]
    sheet.cell(row=53, column=4).value = float(feecodeList[38][3]) * typcasted_dict[
        'Dry Matter'] / 100  # dry matter division and multiplication
    sheet.cell(row=54, column=4).value = float(feecodeList[39][3]) * typcasted_dict[
        'Dry Matter'] / 100  # dry matter division

    # Agricultural End-Use 1
    sheet.cell(row=59, column=4).value = feecodeList[10][3]
    sheet.cell(row=60, column=4).value = feecodeList[8][3]
    sheet.cell(row=61, column=4).value = feecodeList[5][3]
    sheet.cell(row=62, column=4).value = feecodeList[9][3]
    sheet.cell(row=63, column=4).value = feecodeList[44][3]
    sheet.cell(row=64, column=4).value = feecodeList[7][3]
    sheet.cell(row=65, column=4).value = feecodeList[45][3]
    sheet.cell(row=66, column=4).value = feecodeList[6][3]

    calc_value = typcasted_dict['Nitrogen Total (N)'] * dry_matter / 100
    sheet.cell(row=70, column=4).value = round(calc_value, 1)

    sheet.cell(row=71, column=4).value = feecodeList[11][3]

    calc_value = typcasted_dict['Nitrate Nitrogen NO3-N'] * dry_matter / 100
    sheet.cell(row=72, column=4).value = calc_value

    calc_value = typcasted_dict['Total Phosphate (P as P2O5)'] * dry_matter / 100
    num_x = calc_value / 10000 * 2.29  # LOOKS SAME BUT IS ACTUALLY DIFFERENT
    sheet.cell(row=73, column=4).value = num_x

    calc_value = typcasted_dict['Total Potash (K as K2O)'] * dry_matter / 100
    num_x = calc_value / 10000 * 1.21  # LOOKS SAME BUT IS ACTUALLY DIFFERENT
    sheet.cell(row=74, column=4).value = num_x

    calc_value = typcasted_dict['Available Sodium (Na)'] * dry_matter / 100 / 10000
    sheet.cell(row=75, column=4).value = calc_value

    calc_value = typcasted_dict['Sodium'] * dry_matter / 100 / 10000
    sheet.cell(row=76, column=4).value = calc_value

    calc_value = typcasted_dict['Total Available (Mg)'] * typcasted_dict['Dry Matter'] / 100 / 10000
    sheet.cell(row=77, column=4).value = calc_value

    calc_value = typcasted_dict['Total Magnesium (Mg)'] * typcasted_dict[
        'Dry Matter'] / 100  # oNLY WORKS IF I DONT DO THE SECOND DIVISION?
    sheet.cell(row=78, column=4).value = calc_value

    calc_value = typcasted_dict['Total available (Ca)'] * typcasted_dict['Dry Matter'] / 100 / 10000
    sheet.cell(row=79, column=4).value = calc_value

    calc_value = typcasted_dict['Total Calcium (Ca)'] * typcasted_dict[
        'Dry Matter'] / 100  # ONLY WOKRS IF I DONT DO SECOND DIVISON
    sheet.cell(row=80, column=4).value = calc_value

    calc_value = typcasted_dict['Available (S}'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=81, column=4).value = round(calc_value, 1)

    calc_value = typcasted_dict['Total Sulfur (S)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=82, column=4).value = calc_value

    calc_value = typcasted_dict['Boron (B)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=83, column=4).value = calc_value

    calc_value = typcasted_dict['Chloride'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=84, column=4).value = calc_value

    calc_value = typcasted_dict['Copper (Cu)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=85, column=4).value = calc_value

    calc_value = typcasted_dict['Iron (Fe)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=86, column=4).value = calc_value

    calc_value = typcasted_dict['Manganese (Mn)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=87, column=4).value = calc_value

    calc_value = typcasted_dict['Molybdenum (Mb)'] * typcasted_dict['Dry Matter'] / 100 / 10000  # small but seems right
    sheet.cell(row=88, column=4).value = calc_value

    calc_value = typcasted_dict['Zinc (Zn)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=89, column=4).value = calc_value

    # -----------------------------------------------------------------
    from Utilities import FixFormatting
    # BORDER ALIGNMENT
    from CQAUtilities import DQA_CFIA_FORMATTING, DQA_CFIA_PERCENT_ADDITION, DQA_CFIA_ADDING_NA
    DQA_CFIA_FORMATTING(sheet)
    DQA_CFIA_PERCENT_ADDITION(sheet)
    DQA_CFIA_ADDING_NA(sheet)

    # CENTERING THINGS
    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    for i in range(1, 100):
        # print(sheet.cell(row=i, column=4).value)
        # sheet.cell(row=i, column=4).value.alignment = Alignment(horizontal='center')
        current_cell = sheet['D%d' % i]
        current_cell.alignment = Alignment(horizontal='center',vertical='center')
        current_cell.font = Font(bold=True, name='Franklin Gothic Book')


    # putting in the images------------------------------------
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')
    img = Image('al.jpg')
    sheet.add_image(img, 'A1')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'J1')

    img = Image('al.jpg')
    sheet.add_image(img, 'A47')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'J47')
    # ----------------------------------------------------------------
    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport",
                                CQARef)
    Workbook.save(saveLocation + "\%sReport.xlsx" % (CQARef))
    pass
