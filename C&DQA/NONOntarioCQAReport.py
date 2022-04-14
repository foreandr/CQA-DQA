import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Side
import gettingFeeCodes
import CQAUtilities
import Utilities

def BCandOtherReport(workbook, CQAREF):
    print 'BCandOtherReport', workbook, CQAREF
    grey_highlight = PatternFill(start_color='E6E6E3', end_color='E6E6E3', fill_type='solid')
    sheet = workbook.get_sheet_by_name("CCME (Provinces&Territories)")
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    pe_m3_dict = CQAUtilities.getOtherResults(CQAREF)
    array_values, soil_result, env_resut = CQAUtilities.OntarioResults(CQAREF)

    Utilities.round_all_array_values(array_values)
    #print('soil result', soil_result)
    #print('env_resut', env_resut)
    print('GOT ALL VALUES\n')
    for i in array_values:
        print(i)
    # locations_in_excel = Utilities.get_names_and_indexes(sheet)

    # PHOSPHORUS
    PH = Utilities.getPH(CQAREF)

    # A. ---
    sheet.cell(row=10, column=4).value = array_values[0][2]  # Arsenic
    sheet.cell(row=11, column=4).value = array_values[1][2]  # Cadmium
    sheet.cell(row=12, column=4).value = array_values[2][2]  # chromium  # cobalt
    sheet.cell(row=13, column=4).value = array_values[3][2]
    sheet.cell(row=14, column=4).value = array_values[4][2]
    sheet.cell(row=15, column=4).value = array_values[5][2]
    sheet.cell(row=16, column=4).value = array_values[6][2]
    sheet.cell(row=17, column=4).value = array_values[7][2]
    sheet.cell(row=18, column=4).value = array_values[8][2]
    sheet.cell(row=19, column=4).value = array_values[9][2]
    sheet.cell(row=20, column=4).value = array_values[10][2]

    # B. foreign matter
    sheet.cell(row=26, column=4).value = array_values[12][2]

    # C. Sharp Foreign Matter
    sheet.cell(row=29, column=4).value = array_values[29][2]
    sheet.cell(row=30, column=4).value = array_values[13][2]

    # D. MATURITY
    sheet.cell(row=34, column=4).value = array_values[14][2]
    sheet.cell(row=36, column=4).value = ""

    # D. Pathogens
    sheet.cell(row=41, column=4).value = Utilities.get_fecal(CQAREF)
    sheet.cell(row=42, column=4).value = array_values[16][2]

    # E. CFIA
    sheet.cell(row=47, column=6).value = str(array_values[17][2]) + '%'
    sheet.cell(row=48, column=6).value = str(array_values[18][2]) + '%'
    sheet.cell(row=55, column=6).value = PH
    sheet.cell(row=56, column=6).value = array_values[20][2]
    sheet.cell(row=57, column=6).value = CQAUtilities.get_partcile(CQAREF)
    sheet.cell(row=58, column=6).value = pe_m3_dict['salt'] # salt
    sheet.cell(row=59, column=6).value = str(pe_m3_dict['perna_m3']) + '%' #perna
    sheet.cell(row=61, column=6).value = str(pe_m3_dict['perk_m3'])  + '%' #perk
    sheet.cell(row=62, column=6).value = str(pe_m3_dict['permg_m3']) + '%' #perma
    sheet.cell(row=63, column=6).value = str(pe_m3_dict['perca_m3']) + '%' #perca


    # Appendix III
    Nitrogen = float(Utilities.getNitrogen(CQAREF))
    print('NITROGEN' ,Nitrogen)
    print('rOUNDED NITROGEN', round(Nitrogen, 2))
    sheet.cell(row= 98, column=4).value = str(CQAUtilities.get_dry_matter(CQAREF)) + '%'
    sheet.cell(row= 99, column=4).value = PH  # TENTATIVE, MAY BE WRONG
    sheet.cell(row=100, column=4).value = env_resut['30']
    sheet.cell(row=101, column=4).value = array_values[20][2]
    sheet.cell(row=103, column=4).value = str(round(Nitrogen, 2)) + '%'
    sheet.cell(row=104, column=4).value = array_values[23][2]
    sheet.cell(row=105, column=4).value = str(array_values[24][2]) + '%'
    sheet.cell(row=106, column=4).value = str(array_values[25][2]) + '%'
    sheet.cell(row=107, column=4).value = str(array_values[26][2]) + '%'
    sheet.cell(row=108, column=4).value = str(array_values[27][2]) + '%'
    sheet.cell(row=109, column=4).value = array_values[28][2]

    # AGINDEX----------------------------------------------------------
    item_dict = Utilities.getValuesForAGIndex(CQAREF)

    print('getting agindex values')
    DryMatter = float(CQAUtilities.get_dry_matter(CQAREF))
    Nitrogen = float(Utilities.getNitrogen(CQAREF))  # stand in for real value
    Phosphorus = (float(CQAUtilities.get_Agindex_Phosphorus(CQAREF) * (DryMatter / 100) / 10000)) * 2.2914
    Potassium = (float(CQAUtilities.get_Agindex_Potassium(CQAREF) * (DryMatter / 100) / 10000)) * 1.2046
    Sodium = float(CQAUtilities.get_Agindex_Sodium(CQAREF) * (DryMatter / 100))
    Chloride = float(item_dict['CL']) / 10000
    # print('dude wtf', Chloride, type(Chloride), type(float(Chloride)))

    print('PHOSPHORUS:', Phosphorus)  # *
    print('Potassium:', Potassium)  # *
    print('Nitrogen:', Nitrogen)
    print('sodium:', Sodium)  # *
    print('DryMatter:', DryMatter)
    print('Chloride:', Chloride)
    value1 = (Nitrogen + Phosphorus + Potassium)
    value2 = Sodium + (Chloride)
    print 'LEFTSIDE:', value1
    print 'RIGHTSIDE:', value2
    ag_index = value1 / value2
    print('AGINDEX = ', ag_index)
    sheet.cell(row=111, column=4).value = round(ag_index, 2)
    sheet.cell(row=111, column=6).value = CQAUtilities.agindex_text(ag_index)

    CQAUtilities.CQA_OTHER_FORMATTING(sheet)

    # -- Agindex
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')
    # ag_index_jpg = Image('C:\CQA\FULL CQA - DQA\C&DQA\Photos\Agindex.jpg')
    ag_index_png = Image('C:/CQA\FULL CQA - DQA/C&DQA/Photos/agindex.png')
    sheet.add_image(ag_index_png, 'A113')
    # ---------------------------------------
    # HIGHLIGHTING
    import HighlighterChecker
    HighlighterChecker.get_non_ontario_cqa_constraints(sheet)

    # putting in the images------------------------------------
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')
    img = Image('al.jpg')
    sheet.add_image(img, 'A1')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'H1')

    img = Image('al.jpg')
    sheet.add_image(img, 'A93')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'H93')

    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport", CQAREF)
    filename = saveLocation + "\%sReport.xlsx" % CQAREF
    Workbook.save(workbook, filename)

