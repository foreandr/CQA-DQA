import os

from openpyxl.styles import PatternFill, Border, Side
from openpyxl import Workbook
import Colors
import Utilities


def OntarioPrintDQA(Workbook, CQARef):
    # Sets the color of the highlight/fill to highlight the failed values
    highlight = PatternFill(start_color='F3F315', end_color='F3F315', fill_type='solid')
    grey_highlight = PatternFill(start_color='E6E6E3', end_color='E6E6E3', fill_type='solid')
    sheet = Workbook.get_sheet_by_name("Ontario CFIA", )
    thick = Side(border_style="medium")
    thin = Side(border_style="thin")

    import gettingFeeCodes

    feecodeList = gettingFeeCodes.gettingfeeCodes('%s' % CQARef)
    print Colors.bcolors.OKGREEN, '\nFEECODELIST', Colors.bcolors.ENDC
    for i in feecodeList:
        print i
    # Some missing, but actually turned out pretty good
    locations_in_excel = Utilities.grab_excel_locations()
    for row in locations_in_excel:
        # print row
        current_index = row[0]
        current_name = row[1]
        for i in feecodeList:
            if i[1] == current_name:
                # print current_index, current_name, i[1]
                # print 'yes'
                sheet.cell(row=current_index, column=4).value = i[3]

    dict_names_values = {}
    for i in feecodeList:
        dict_names_values[i[1]] = (i[3])
    typcasted_dict = {}
    print '\n'
    for key, value in dict_names_values.iteritems():
        try:
            typcasted_dict[key] = float(value)
        except:
            # print(value, 'not a float, entering into dict as string')
            typcasted_dict[key] = value
    # A
    sheet.cell(row=17, column=4).value = typcasted_dict['Molybdenum (Mb)']

    # B
    sheet.cell(row=28, column=4).value = feecodeList[36][3]
    sheet.cell(row=29, column=4).value = feecodeList[37][3]
    sheet.cell(row=30, column=4).value = feecodeList[35][3]

    # C Pathogens
    sheet.cell(row=36, column=4).value = feecodeList[1][3]
    sheet.cell(row=35, column=4).value = feecodeList[3][3]

    # E Physical Parameter
    # sheet.cell(row=53, column=4).value = feecodeList[7][3]
    # sheet.cell(row=54, column=4).value = feecodeList[45][3]
    # sheet.cell(row=55, column=4).value = feecodeList[6][3]

    # Minimum Agricultural Values
    sheet.cell(row=53, column=4).value = feecodeList[11][3]
    sheet.cell(row=54, column=4).value = float(feecodeList[38][3]) * typcasted_dict[
        'Dry Matter'] / 100  # this is total phophate not available
    sheet.cell(row=55, column=4).value = float(feecodeList[39][3]) * typcasted_dict['Dry Matter'] / 100

    # Agricultural End-Use
    sheet.cell(row=60, column=4).value = str(sheet.cell(row=60, column=4).value) + '%'
    sheet.cell(row=63, column=4).value = feecodeList[9][3]
    sheet.cell(row=64, column=4).value = feecodeList[44][3]  # can't seem to find
    sheet.cell(row=65, column=4).value = Utilities.getTotalOrganicMatter(CQARef)
    sheet.cell(row=66, column=4).value = float(feecodeList[45][3])
    sheet.cell(row=67, column=4).value = str(feecodeList[6][3]) + '%'

    # Fertilizer Equivalent Materials

    sheet.cell(row=76, column=4).value = feecodeList[28][3]  # SODIUM

    # ----------------------------------------------------------------    print '\n'
    # for key, value in typcasted_dict.iteritems():
    #   print key, value

    '''FOR MULTIPLICATION PURPOSES'''
    # print('\nDRY MATTER : ')
    # print(typcasted_dict['Dry Matter'])
    calc_value = typcasted_dict['Nitrogen Total (N)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=71, column=4).value = round(calc_value, 1)

    calc_value = typcasted_dict['Nitrate Nitrogen NO3-N'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=73, column=4).value = calc_value

    calc_value = typcasted_dict['Total Phosphate (P as P2O5)'] * typcasted_dict['Dry Matter'] / 100
    num_x = calc_value / 10000 * 2.29  # LOOKS SAME BUT IS ACTUALLY DIFFERENT
    sheet.cell(row=74, column=4).value = num_x

    calc_value = typcasted_dict['Total Potash (K as K2O)'] * typcasted_dict['Dry Matter'] / 100
    num_x = calc_value / 10000 * 1.21  # LOOKS SAME BUT IS ACTUALLY DIFFERENT
    sheet.cell(row=75, column=4).value = num_x

    calc_value = typcasted_dict['Available Sodium (Na)'] * typcasted_dict['Dry Matter'] / 100 / 10000
    # print('sodium calc value', calc_value)
    sheet.cell(row=76, column=4).value = calc_value

    calc_value = typcasted_dict['Sodium'] * typcasted_dict['Dry Matter'] / 100 / 10000
    sheet.cell(row=77, column=4).value = calc_value

    calc_value = typcasted_dict['Total Available (Mg)'] * typcasted_dict['Dry Matter'] / 100 / 10000
    sheet.cell(row=78, column=4).value = calc_value

    calc_value = typcasted_dict['Total Magnesium (Mg)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=79, column=4).value = calc_value

    calc_value = typcasted_dict['Total available (Ca)'] * typcasted_dict['Dry Matter'] / 100 / 10000
    sheet.cell(row=80, column=4).value = calc_value

    calc_value = typcasted_dict['Total Calcium (Ca)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=81, column=4).value = calc_value

    calc_value = typcasted_dict['Available (S}'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=82, column=4).value = round(calc_value, 1)

    calc_value = typcasted_dict['Total Sulfur (S)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=83, column=4).value = calc_value

    calc_value = typcasted_dict['Boron (B)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=84, column=4).value = calc_value

    calc_value = typcasted_dict['Chloride'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=85, column=4).value = calc_value

    calc_value = typcasted_dict['Copper (Cu)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=86, column=4).value = calc_value

    calc_value = typcasted_dict['Iron (Fe)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=87, column=4).value = calc_value

    calc_value = typcasted_dict['Manganese (Mn)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=88, column=4).value = calc_value

    calc_value = typcasted_dict['Molybdenum (Mb)'] * typcasted_dict['Dry Matter'] / 100 / 10000  # small but seems right
    sheet.cell(row=89, column=4).value = calc_value

    calc_value = typcasted_dict['Zinc (Zn)'] * typcasted_dict['Dry Matter'] / 100
    sheet.cell(row=90, column=4).value = calc_value
    #-------
    sheet.cell(row=71, column=4).value = str(round(float(sheet.cell(row=71, column=4).value), 2)) + '%'
    sheet.cell(row=74, column=4).value = str(round(float(sheet.cell(row=74, column=4).value), 2)) + '%'
    sheet.cell(row=75, column=4).value = str(round(float(sheet.cell(row=75, column=4).value), 2)) + '%'
    sheet.cell(row=76, column=4).value = str(round(float(sheet.cell(row=76, column=4).value), 2)) + '%'
    sheet.cell(row=77, column=4).value = str(round(float(sheet.cell(row=77, column=4).value), 2)) + '%'
    sheet.cell(row=78, column=4).value = str(round(float(sheet.cell(row=78, column=4).value), 2)) + '%'
    sheet.cell(row=79, column=4).value = str(round(float(sheet.cell(row=79, column=4).value), 2)) + '%'
    sheet.cell(row=80, column=4).value = str(round(float(sheet.cell(row=80, column=4).value), 2)) + '%'
    sheet.cell(row=81, column=4).value = str(round(float(sheet.cell(row=81, column=4).value), 2)) + '%'
    # --------------------------------------------
    from CQAUtilities import DQA_ONT_FORMATTING
    DQA_ONT_FORMATTING(sheet)

    Utilities.number_formatting(sheet)
    # -----------------------------------------------------------------
    # BORDER ALIGNMENT

    # CENTERING THINGS

    from openpyxl.styles import Alignment
    from openpyxl.styles import Font
    for i in range(1, 100):
        # print(sheet.cell(row=i, column=4).value)
        # sheet.cell(row=i, column=4).value.alignment = Alignment(horizontal='center')
        current_cell = sheet['D%d' % i]
        current_cell.alignment = Alignment(horizontal='center')
        current_cell.font = Font(bold=True, name='Franklin Gothic Book')

        if i in range(70, 90):
            current_cell = sheet['F%d' % i]
            current_cell.alignment = Alignment(horizontal='center')
            current_cell = sheet['G%d' % i]
            current_cell.alignment = Alignment(horizontal='center')
            current_cell = sheet['H%d' % i]
            current_cell.alignment = Alignment(horizontal='center')
            current_cell = sheet['I%d' % i]
            current_cell.alignment = Alignment(horizontal='center')

    if float(sheet['D88'].value) < 1.0:  # should proably turn this into a function, way easier and more maliable
        sheet['D88'] = 'BDL'
        sheet['F88'] = 'N/A'
        sheet['G88'] = 'N/A'
        sheet['H88'] = 'N/A'
        sheet['I88'] = 'N/A'

    # putting in the images------------------------------------
    from openpyxl.drawing.image import Image
    os.chdir(r'C:\CQA\FULL CQA - DQA\C&DQA\Photos')

    img = Image('al.jpg')
    sheet.add_image(img, 'B1')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'I1')

    img = Image('al.jpg')
    sheet.add_image(img, 'A48')
    img = Image('Digestate-logo.png')
    sheet.add_image(img, 'I48')

    # ----------------------------------------------------------------
    saveLocation = os.path.join(r"C:\CQA\FULL CQA - DQA\C&DQA\FinishedReport",
                                CQARef)
    Workbook.save(saveLocation + "\%sReport.xlsx" % (CQARef))
